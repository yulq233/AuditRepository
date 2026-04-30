"""
导出服务
支持导出抽凭结果、工作底稿等到 Excel/PDF
"""
import io
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
import uuid

from app.core.database import get_db_cursor
from app.core.config import settings


class ExcelExporter:
    """Excel 导出器"""

    def __init__(self):
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            self.pd = pd
            self.Workbook = Workbook
            self.styles = {
                'Font': Font,
                'Alignment': Alignment,
                'Border': Border,
                'Side': Side,
                'PatternFill': PatternFill
            }
            self.available = True
        except ImportError:
            self.available = False

    def export_sampling_results(
        self,
        project_id: str,
        rule_id: Optional[str] = None,
        include_details: bool = True
    ) -> bytes:
        """
        导出抽凭结果到 Excel

        Args:
            project_id: 项目 ID
            rule_id: 规则 ID（可选，不指定则导出全部）
            include_details: 是否包含详细信息

        Returns:
            Excel 文件字节
        """
        if not self.available:
            raise RuntimeError("请安装 pandas 和 openpyxl: pip install pandas openpyxl")

        # 查询数据
        with get_db_cursor() as cursor:
            # 获取项目信息
            cursor.execute(
                "SELECT name FROM projects WHERE id = ?",
                [project_id]
            )
            project_row = cursor.fetchone()
            project_name = project_row[0] if project_row else "未知项目"

            # 构建查询
            query = """
                SELECT
                    v.voucher_no,
                    v.voucher_date,
                    v.amount,
                    v.subject_code,
                    v.subject_name,
                    v.description,
                    v.counterparty,
                    s.risk_score,
                    s.reason,
                    sr.name as rule_name,
                    s.sampled_at
                FROM samples s
                JOIN vouchers v ON s.voucher_id = v.id
                LEFT JOIN sampling_rules sr ON s.rule_id = sr.id
                WHERE s.project_id = ?
            """
            params = [project_id]

            if rule_id:
                query += " AND s.rule_id = ?"
                params.append(rule_id)

            query += " ORDER BY s.sampled_at DESC"

            cursor.execute(query, params)
            rows = cursor.fetchall()

        # 创建 DataFrame
        columns = [
            "凭证编号", "凭证日期", "金额", "科目代码",
            "科目名称", "摘要", "交易对手", "风险分数",
            "抽取原因", "规则名称", "抽取时间"
        ]

        data = []
        for row in rows:
            data.append({
                "凭证编号": row[0],
                "凭证日期": str(row[1]) if row[1] else "",
                "金额": float(row[2]) if row[2] else 0,
                "科目代码": row[3] or "",
                "科目名称": row[4] or "",
                "摘要": row[5] or "",
                "交易对手": row[6] or "",
                "风险分数": float(row[7]) if row[7] else "",
                "抽取原因": row[8] or "",
                "规则名称": row[9] or "",
                "抽取时间": str(row[10]) if row[10] else ""
            })

        df = self.pd.DataFrame(data)

        # 导出到 Excel
        output = io.BytesIO()

        with self.pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='抽凭结果', index=False)

            # 获取工作簿和工作表
            workbook = writer.book
            worksheet = writer.sheets['抽凭结果']

            # 设置样式
            header_font = self.styles['Font'](bold=True, size=11)
            header_fill = self.styles['PatternFill'](
                start_color='4472C4',
                end_color='4472C4',
                fill_type='solid'
            )
            header_font_white = self.styles['Font'](bold=True, color='FFFFFF', size=11)
            thin_border = self.styles['Border'](
                left=self.styles['Side'](style='thin'),
                right=self.styles['Side'](style='thin'),
                top=self.styles['Side'](style='thin'),
                bottom=self.styles['Side'](style='thin')
            )

            # 设置表头样式
            for cell in worksheet[1]:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = self.styles['Alignment'](horizontal='center', vertical='center')
                cell.border = thin_border

            # 设置数据区域样式
            for row in worksheet.iter_rows(min_row=2, max_row=len(data) + 1):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = self.styles['Alignment'](vertical='center')

            # 调整列宽
            column_widths = {
                'A': 15,  # 凭证编号
                'B': 12,  # 凭证日期
                'C': 12,  # 金额
                'D': 12,  # 科目代码
                'E': 20,  # 科目名称
                'F': 30,  # 摘要
                'G': 20,  # 交易对手
                'H': 10,  # 风险分数
                'I': 25,  # 抽取原因
                'J': 15,  # 规则名称
                'K': 20,  # 抽取时间
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

        output.seek(0)
        return output.getvalue()

    def export_vouchers(
        self,
        project_id: str,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        导出凭证列表到 Excel

        Args:
            project_id: 项目 ID
            filters: 筛选条件

        Returns:
            Excel 文件字节
        """
        if not self.available:
            raise RuntimeError("请安装 pandas 和 openpyxl: pip install pandas openpyxl")

        # 查询数据
        with get_db_cursor() as cursor:
            query = """
                SELECT
                    voucher_no, voucher_date, amount,
                    subject_code, subject_name, description,
                    counterparty, created_at
                FROM vouchers
                WHERE project_id = ?
            """
            params = [project_id]

            # 应用筛选条件
            if filters:
                if filters.get('subject_code'):
                    query += " AND subject_code = ?"
                    params.append(filters['subject_code'])
                if filters.get('start_date'):
                    query += " AND voucher_date >= ?"
                    params.append(filters['start_date'])
                if filters.get('end_date'):
                    query += " AND voucher_date <= ?"
                    params.append(filters['end_date'])

            query += " ORDER BY voucher_date, voucher_no"

            cursor.execute(query, params)
            rows = cursor.fetchall()

        # 创建 DataFrame
        data = []
        for row in rows:
            data.append({
                "凭证编号": row[0],
                "凭证日期": str(row[1]) if row[1] else "",
                "金额": float(row[2]) if row[2] else 0,
                "科目代码": row[3] or "",
                "科目名称": row[4] or "",
                "摘要": row[5] or "",
                "交易对手": row[6] or "",
                "创建时间": str(row[7]) if row[7] else ""
            })

        df = self.pd.DataFrame(data)

        # 导出到 Excel
        output = io.BytesIO()
        df.to_excel(output, index=False, sheet_name='凭证列表')
        output.seek(0)

        return output.getvalue()

    def export_project_summary(
        self,
        project_id: str
    ) -> bytes:
        """
        导出项目汇总报告

        Args:
            project_id: 项目 ID

        Returns:
            Excel 文件字节
        """
        if not self.available:
            raise RuntimeError("请安装 pandas 和 openpyxl: pip install pandas openpyxl")

        output = io.BytesIO()

        with self.pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. 项目概况
            with get_db_cursor() as cursor:
                cursor.execute(
                    "SELECT name, description, status, created_at FROM projects WHERE id = ?",
                    [project_id]
                )
                project = cursor.fetchone()

                cursor.execute(
                    "SELECT COUNT(*), COALESCE(SUM(amount), 0) FROM vouchers WHERE project_id = ?",
                    [project_id]
                )
                voucher_stats = cursor.fetchone()

                cursor.execute(
                    "SELECT COUNT(*) FROM samples WHERE project_id = ?",
                    [project_id]
                )
                sample_count = cursor.fetchone()[0]

            overview_data = {
                "项目名称": [project[0]],
                "项目描述": [project[1] or ""],
                "项目状态": [project[2]],
                "创建时间": [str(project[3])],
                "凭证总数": [voucher_stats[0]],
                "凭证总金额": [float(voucher_stats[1])],
                "抽样数量": [sample_count]
            }

            df_overview = self.pd.DataFrame(overview_data)
            df_overview.to_excel(writer, sheet_name='项目概况', index=False)

            # 2. 科目分布
            with get_db_cursor() as cursor:
                cursor.execute(
                    """
                    SELECT subject_name, COUNT(*) as cnt, SUM(amount) as total
                    FROM vouchers
                    WHERE project_id = ? AND subject_name IS NOT NULL
                    GROUP BY subject_name
                    ORDER BY total DESC
                    """,
                    [project_id]
                )
                subject_rows = cursor.fetchall()

            subject_data = []
            for row in subject_rows:
                subject_data.append({
                    "科目名称": row[0],
                    "凭证数量": row[1],
                    "金额合计": float(row[2])
                })

            df_subject = self.pd.DataFrame(subject_data)
            df_subject.to_excel(writer, sheet_name='科目分布', index=False)

            # 3. 抽样结果
            with get_db_cursor() as cursor:
                cursor.execute(
                    """
                    SELECT v.voucher_no, v.voucher_date, v.amount,
                           v.subject_name, v.description, s.risk_score, s.reason
                    FROM samples s
                    JOIN vouchers v ON s.voucher_id = v.id
                    WHERE s.project_id = ?
                    """,
                    [project_id]
                )
                sample_rows = cursor.fetchall()

            sample_data = []
            for row in sample_rows:
                sample_data.append({
                    "凭证编号": row[0],
                    "日期": str(row[1]) if row[1] else "",
                    "金额": float(row[2]) if row[2] else 0,
                    "科目": row[3] or "",
                    "摘要": row[4] or "",
                    "风险分": float(row[5]) if row[5] else "",
                    "抽取原因": row[6] or ""
                })

            df_samples = self.pd.DataFrame(sample_data)
            df_samples.to_excel(writer, sheet_name='抽样结果', index=False)

        output.seek(0)
        return output.getvalue()


class PDFExporter:
    """PDF 导出器"""

    def __init__(self):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            self.colors = colors
            self.A4 = A4
            self.SimpleDocTemplate = SimpleDocTemplate
            self.Table = Table
            self.TableStyle = TableStyle
            self.Paragraph = Paragraph
            self.Spacer = Spacer
            self.getSampleStyleSheet = getSampleStyleSheet
            self.available = True

            # 尝试注册中文字体
            try:
                # 尝试常见中文字体路径
                import os
                font_paths = [
                    'C:/Windows/Fonts/simhei.ttf',  # Windows 黑体
                    'C:/Windows/Fonts/simsun.ttc',  # Windows 宋体
                    'C:/Windows/Fonts/msyh.ttc',    # Windows 微软雅黑
                    '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',  # Linux 文泉驿
                    '/System/Library/Fonts/PingFang.ttc',  # macOS 苹方
                ]
                for font_path in font_paths:
                    try:
                        if os.path.exists(font_path):
                            pdfmetrics.registerFont(TTFont('Chinese', font_path))
                            self.chinese_font = 'Chinese'
                            break
                    except:
                        continue
                else:
                    self.chinese_font = None
            except:
                self.chinese_font = None

        except ImportError:
            self.available = False

    def export_working_paper(
        self,
        paper: Dict[str, Any],
        project_info: Dict[str, Any] = None
    ) -> bytes:
        """
        导出工作底稿到 PDF

        Args:
            paper: 底稿数据（包含 title, sections, ai_description）
            project_info: 项目信息

        Returns:
            PDF 文件字节
        """
        if not self.available:
            raise RuntimeError("请安装 reportlab: pip install reportlab")

        output = io.BytesIO()

        doc = self.SimpleDocTemplate(output, pagesize=self.A4)
        elements = []

        # 样式
        styles = self.getSampleStyleSheet()
        if self.chinese_font:
            title_style = styles['Heading1']
            title_style.fontName = self.chinese_font
            normal_style = styles['Normal']
            normal_style.fontName = self.chinese_font
            heading2_style = styles['Heading2']
            heading2_style.fontName = self.chinese_font
        else:
            normal_style = styles['Normal']
            heading2_style = styles['Heading2']

        # 标题
        title = paper.get('title', '工作底稿')
        elements.append(self.Paragraph(title, styles['Heading1']))
        elements.append(self.Spacer(1, 10))

        # 项目信息
        if project_info:
            info_data = [
                ["项目名称", project_info.get('name', '')],
                ["底稿类型", paper.get('paper_type', '')],
            ]

            info_table = self.Table(info_data, colWidths=[100, 300])
            info_table.setStyle(self.TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), self.chinese_font or 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ]))
            elements.append(info_table)
            elements.append(self.Spacer(1, 15))

        # 底稿章节内容
        sections = paper.get('sections', [])
        for section in sections:
            section_title = section.get('title', '')
            section_content = section.get('content', '')

            # 章节标题
            elements.append(self.Paragraph(section_title, heading2_style))
            elements.append(self.Spacer(1, 5))

            # 章节内容（处理多行文本）
            if section_content:
                # 将内容按换行符分割，每段作为一个 Paragraph
                paragraphs = section_content.split('\n')
                for para in paragraphs:
                    if para.strip():
                        elements.append(self.Paragraph(para.strip(), normal_style))
                        elements.append(self.Spacer(1, 2))
                elements.append(self.Spacer(1, 10))

        # AI 描述
        ai_description = paper.get('ai_description', '')
        if ai_description:
            elements.append(self.Paragraph("AI 审计说明", heading2_style))
            elements.append(self.Spacer(1, 5))

            ai_paragraphs = ai_description.split('\n')
            for para in ai_paragraphs:
                if para.strip():
                    elements.append(self.Paragraph(para.strip(), normal_style))
                    elements.append(self.Spacer(1, 2))

        # 构建 PDF
        doc.build(elements)
        output.seek(0)

        return output.getvalue()


# 全局导出器实例
excel_exporter = ExcelExporter()
pdf_exporter = PDFExporter()


def save_export_file(
    file_data: bytes,
    filename: str,
    project_id: str = None
) -> str:
    """
    保存导出文件

    Args:
        file_data: 文件数据
        filename: 文件名
        project_id: 项目 ID

    Returns:
        文件路径
    """
    # 创建导出目录
    export_dir = os.path.join(settings.PAPER_PATH, project_id or "")
    os.makedirs(export_dir, exist_ok=True)

    # 生成文件路径
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    full_filename = f"{timestamp}_{filename}"
    file_path = os.path.join(export_dir, full_filename)

    # 保存文件
    with open(file_path, 'wb') as f:
        f.write(file_data)

    return file_path
