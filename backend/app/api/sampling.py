"""
智能抽样API
"""
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Form, Depends
from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any
from datetime import datetime
import uuid
import json
import os
import shutil
import logging
import time

logger = logging.getLogger(__name__)

from app.core.database import get_db_cursor, get_db
from app.core.config import settings
from app.core.auth import get_current_user, UserInDB
from app.services.ai_sampling_service import risk_analyzer
from app.services.voucher_risk_service import voucher_risk_service
import asyncio

router = APIRouter()


# ==================== 数据模型 ====================

class SamplingRuleCreate(BaseModel):
    """创建抽凭规则"""
    name: str = Field(..., description="规则名称")
    rule_type: str = Field(..., description="规则类型: random/amount/subject/date/stratified/ai")
    rule_config: dict = Field(..., description="规则配置")


class StratifiedLayerConfig(BaseModel):
    """分层抽样配置 - 单层"""
    name: str = Field(..., description="层名称")
    min_amount: Optional[float] = Field(None, alias="minAmount", description="最小金额")
    max_amount: Optional[float] = Field(None, alias="maxAmount", description="最大金额")
    sampling_rate: float = Field(..., alias="samplingRate", description="抽样比例 (0-1)")

    class Config:
        populate_by_name = True  # 允许通过字段名或别名填充


class StratifiedSamplingConfig(BaseModel):
    """分层抽样配置"""
    stratify_by: str = Field(default="amount", description="分层依据: amount(金额)")
    layers: List[StratifiedLayerConfig] = Field(..., description="分层配置")


class SamplingRuleResponse(BaseModel):
    """抽凭规则响应"""
    id: str
    project_id: str
    name: str
    rule_type: str
    rule_config: dict
    is_active: bool
    created_at: datetime


class SamplingExecuteRequest(BaseModel):
    """执行抽凭请求"""
    rule_id: str = Field(..., description="规则ID")
    sample_size: Optional[int] = Field(None, description="样本数量")


def calculate_risk_details(risk_score: Optional[float], amount: Optional[float], voucher_data: dict = None) -> dict:
    """根据风险分数和金额计算风险详情（降级方案）

    Args:
        risk_score: 风险分数
        amount: 金额
        voucher_data: 凭证数据（用于更精确的分析）
    """
    if risk_score is None:
        risk_score = 30  # 默认分数

    # 构建风险因素列表
    risk_factors = []

    # 1. 金额风险分析
    if amount:
        if amount > 500000:
            risk_factors.append("超大额交易（金额超过50万元），需重点关注")
            risk_score = max(risk_score, 80)
        elif amount > 100000:
            risk_factors.append("大额交易（金额超过10万元）")
            risk_score = max(risk_score, 60)
        elif amount > 50000:
            risk_factors.append("中等金额交易（金额超过5万元）")

        # 检查金额是否为整数（可能的估算值）
        if amount == int(amount) and amount >= 10000:
            risk_factors.append("金额为整数，可能为估算值，需关注业务合理性")

        # 检查金额是否接近整数万（可能的凑整）
        if amount % 10000 == 0 and amount >= 10000:
            risk_factors.append("金额为整数万，可能存在凑整或估算情况")

    # 2. 凭证数据分析
    if voucher_data:
        # 日期风险分析
        voucher_date = voucher_data.get('voucher_date')
        if voucher_date and voucher_date != 'None':
            try:
                from datetime import datetime
                # 解析日期
                if isinstance(voucher_date, str):
                    date_obj = datetime.strptime(voucher_date[:10], '%Y-%m-%d')
                else:
                    date_obj = voucher_date

                # 检查是否为月末（最后3天）
                if date_obj.day >= 28:
                    risk_factors.append(f"月末交易（{voucher_date}），需关注跨期调整风险")
                    risk_score = min(100, risk_score + 10)

                # 检查是否为年末
                if date_obj.month == 12 and date_obj.day >= 25:
                    risk_factors.append(f"年末集中交易（{voucher_date}），可能存在利润调节")
                    risk_score = min(100, risk_score + 15)

                # 检查是否为月初（1-3日）
                if date_obj.day <= 3:
                    risk_factors.append(f"月初集中交易（{voucher_date}），可能存在上月延迟入账")
                    risk_score = min(100, risk_score + 5)

                # 检查是否为周末
                if date_obj.weekday() >= 5:  # 5=周六, 6=周日
                    risk_factors.append(f"非工作日交易（{voucher_date}），需核实业务真实性")
                    risk_score = min(100, risk_score + 10)
            except Exception:
                # 日期解析失败，忽略
                pass

        # 摘要分析
        description = voucher_data.get('description', '')
        if description:
            # 检查摘要中的敏感词
            sensitive_words = ['调整', '冲销', '暂估', '补记', '更正', '结转', '待摊', '预提', '挂账', '清理']
            found_sensitive = [w for w in sensitive_words if w in description]
            if found_sensitive:
                risk_factors.append(f"摘要包含敏感词：{'、'.join(found_sensitive)}")
                risk_score = min(100, risk_score + 15)

            # 检查摘要长度
            if len(description) < 8:
                risk_factors.append("摘要描述过于简短，无法明确业务实质")
            elif len(description) < 15:
                risk_factors.append("摘要描述较简短，建议核实业务内容")
        else:
            risk_factors.append("缺少摘要信息，无法判断业务性质")
            risk_score = min(100, risk_score + 10)

        # 科目风险分析
        subject_name = voucher_data.get('subject_name', '')
        if subject_name:
            # 高风险科目
            high_risk_subjects = ['其他应收款', '其他应付款', '待处理财产损溢', '以前年度损益调整']
            if any(s in subject_name for s in high_risk_subjects):
                risk_factors.append(f"涉及高风险科目：{subject_name}")
                risk_score = min(100, risk_score + 20)

            # 需要关注的科目
            medium_risk_subjects = ['预付账款', '预收账款', '递延所得税', '长期待摊']
            if any(s in subject_name for s in medium_risk_subjects):
                risk_factors.append(f"涉及需关注科目：{subject_name}")
                risk_score = min(100, risk_score + 10)

        # 交易对手分析
        counterparty = voucher_data.get('counterparty', '')
        if counterparty:
            # 关联方关键词
            related_keywords = ['关联', '股东', '子公司', '母公司', '兄弟公司', '联营', '合营']
            if any(k in counterparty for k in related_keywords):
                risk_factors.append(f"疑似关联方交易：{counterparty}")
                risk_score = min(100, risk_score + 25)

            # 个人交易
            if '个人' in counterparty or len(counterparty) <= 3:
                risk_factors.append(f"交易对手为个人或名称简短：{counterparty}")
        else:
            if amount and amount > 50000:
                risk_factors.append("大额交易缺少交易对手信息")

    # 3. 综合风险评分调整
    if risk_score >= 70:
        risk_level = "high"
    elif risk_score >= 40:
        risk_level = "medium"
    else:
        risk_level = "low"

    # 如果没有识别到风险因素，添加默认说明
    if not risk_factors:
        risk_factors.append("常规凭证，各项指标正常")

    # 构建详细的解释说明
    amount_str = f"¥{amount:,.2f}" if amount else "未知"
    voucher_no = voucher_data.get('voucher_no', '') if voucher_data else ''

    if risk_level == "high":
        explanation = f"凭证{voucher_no}风险评分{int(risk_score)}分，属于高风险凭证。"
        explanation += f"金额{amount_str}，主要风险点：{'；'.join(risk_factors[:4])}。"
        explanation += "建议重点审查：1）业务真实性和合规性；2）相关合同、发票等原始凭证的完整性；"
        explanation += "3）交易背景和资金流向；4）是否存在异常交易或舞弊嫌疑。"
    elif risk_level == "medium":
        explanation = f"凭证{voucher_no}风险评分{int(risk_score)}分，属于中风险凭证。"
        explanation += f"金额{amount_str}，需关注：{'；'.join(risk_factors[:4])}。"
        explanation += "建议进一步核实：1）业务背景和交易目的；2）凭证附件是否齐全；"
        explanation += "3）交易的真实性和合规性；4）相关审批流程是否完整。"
    else:
        explanation = f"凭证{voucher_no}风险评分{int(risk_score)}分，属于低风险凭证。金额{amount_str}。"
        explanation += "各项指标基本正常，未发现明显异常迹象。"
        explanation += "可按常规审计程序进行抽查，核实凭证的完整性和准确性。"

    return {
        "risk_level": risk_level,
        "risk_factors": risk_factors,
        "explanation": explanation
    }


async def analyze_vouchers_with_ai(voucher_ids: List[str], project_id: str) -> Dict[str, Dict[str, Any]]:
    """使用AI分析凭证风险（统一使用voucher_risk_service，与风险画像保持一致）

    Args:
        voucher_ids: 凭证ID列表
        project_id: 项目ID

    Returns:
        字典，key为voucher_id，value为风险评估结果
    """
    start_time = time.time()
    logger.info(f"[AI分析] ========== 开始分析（统一服务） ==========")
    logger.info(f"[AI分析] 项目ID: {project_id}")
    logger.info(f"[AI分析] 凭证ID数量: {len(voucher_ids)}")

    try:
        # 使用voucher_risk_service进行AI分析（与风险画像执行分层抽样使用相同服务）
        filters = {"voucher_ids": voucher_ids} if voucher_ids else None
        results_list = await voucher_risk_service.batch_analyze_vouchers(
            project_id=project_id,
            filters=filters,
            use_ai=True
        )

        # 构建结果字典
        results = {}
        for result in results_list:
            results[result.voucher_id] = {
                "risk_score": result.risk_score,
                "risk_level": result.risk_level,
                "risk_factors": result.risk_factors,
                "explanation": result.explanation,
                "confidence": 0.8  # 统一使用默认置信度
            }
            logger.debug(f"[AI分析] 凭证 {result.voucher_id}: 风险={result.risk_level}, 分数={result.risk_score}")

        elapsed = time.time() - start_time
        logger.info(f"[AI分析] ========== 分析成功，共 {len(results)} 条，耗时: {elapsed:.2f}秒 ==========")
        return results

    except Exception as e:
        elapsed = time.time() - start_time
        logger.error(f"[AI分析] ❌ 失败! 耗时: {elapsed:.2f}秒")
        logger.error(f"[AI分析] 错误类型: {type(e).__name__}")
        logger.error(f"[AI分析] 错误详情: {str(e)}")
        import traceback
        logger.error(f"[AI分析] 堆栈跟踪:\n{traceback.format_exc()}")
        return {}


class SampleResponse(BaseModel):
    """抽样结果"""
    id: str
    voucher_id: str
    voucher_no: str
    amount: Optional[float]
    subject_name: Optional[str]
    description: Optional[str]
    risk_score: Optional[float]
    risk_level: Optional[str]
    risk_factors: Optional[List[str]]
    explanation: Optional[str]
    reason: Optional[str]


class SamplingResultResponse(BaseModel):
    """抽凭结果响应"""
    rule_id: str
    rule_name: str
    total_population: int
    sample_size: int
    samples: List[SampleResponse]


class LayerSummary(BaseModel):
    """分层抽样单层统计"""
    name: str
    population_count: int
    sample_count: int
    sampling_rate: float


class StratifiedSamplingResultResponse(BaseModel):
    """分层抽样结果响应"""
    rule_id: str
    rule_name: str
    total_population: int
    sample_size: int
    layers_summary: List[LayerSummary]
    samples: List[SampleResponse]


class RiskProfileResponse(BaseModel):
    """风险画像响应"""
    id: str
    subject_code: str
    subject_name: Optional[str]
    risk_level: str
    risk_factors: List[str]
    material_amount: Optional[float]
    anomaly_score: Optional[float]
    recommendation: Optional[str]


# ==================== API端点 ====================

@router.get("/projects/{project_id}/sampling-rules", response_model=List[SamplingRuleResponse])
async def list_sampling_rules(
    project_id: str,
    current_user: UserInDB = Depends(get_current_user)
):
    """获取抽凭规则列表"""
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        cursor.execute(
            """
            SELECT id, project_id, name, rule_type, rule_config, is_active, created_at
            FROM sampling_rules
            WHERE project_id = ?
            ORDER BY created_at DESC
            """,
            [project_id]
        )
        rows = cursor.fetchall()

        return [
            SamplingRuleResponse(
                id=row[0],
                project_id=row[1],
                name=row[2],
                rule_type=row[3],
                rule_config=json.loads(row[4]) if isinstance(row[4], str) else row[4],
                is_active=bool(row[5]),
                created_at=row[6]
            )
            for row in rows
        ]


@router.post("/projects/{project_id}/sampling-rules", response_model=SamplingRuleResponse, status_code=201)
async def create_sampling_rule(
    project_id: str,
    rule: SamplingRuleCreate,
    current_user: UserInDB = Depends(get_current_user)
):
    """创建抽凭规则"""
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        rule_id = str(uuid.uuid4())
        now = datetime.now()

        cursor.execute(
            """
            INSERT INTO sampling_rules (id, project_id, name, rule_type, rule_config, is_active, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [rule_id, project_id, rule.name, rule.rule_type, json.dumps(rule.rule_config), True, now]
        )
        get_db().commit()

        return SamplingRuleResponse(
            id=rule_id,
            project_id=project_id,
            name=rule.name,
            rule_type=rule.rule_type,
            rule_config=rule.rule_config,
            is_active=True,
            created_at=now
        )


@router.put("/projects/{project_id}/sampling-rules/{rule_id}", response_model=SamplingRuleResponse)
async def update_sampling_rule(
    project_id: str,
    rule_id: str,
    rule: SamplingRuleCreate,
    current_user: UserInDB = Depends(get_current_user)
):
    """更新抽凭规则"""
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        # 验证规则存在
        cursor.execute(
            "SELECT id FROM sampling_rules WHERE id = ? AND project_id = ?",
            [rule_id, project_id]
        )
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="规则不存在")

        # 更新规则
        now = datetime.now()
        cursor.execute(
            """
            UPDATE sampling_rules
            SET name = ?, rule_type = ?, rule_config = ?
            WHERE id = ? AND project_id = ?
            """,
            [rule.name, rule.rule_type, json.dumps(rule.rule_config), rule_id, project_id]
        )
        get_db().commit()

        return SamplingRuleResponse(
            id=rule_id,
            project_id=project_id,
            name=rule.name,
            rule_type=rule.rule_type,
            rule_config=rule.rule_config,
            is_active=True,
            created_at=now
        )


@router.delete("/projects/{project_id}/sampling-rules/{rule_id}")
async def delete_sampling_rule(
    project_id: str,
    rule_id: str,
    current_user: UserInDB = Depends(get_current_user)
):
    """删除抽凭规则"""
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        # 验证规则存在
        cursor.execute(
            "SELECT id FROM sampling_rules WHERE id = ? AND project_id = ?",
            [rule_id, project_id]
        )
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="规则不存在")

        # 删除规则
        cursor.execute("DELETE FROM sampling_rules WHERE id = ?", [rule_id])
        get_db().commit()

        return {"message": "规则删除成功"}


@router.post("/projects/{project_id}/sampling/execute", response_model=SamplingResultResponse)
async def execute_sampling(
    project_id: str,
    request: SamplingExecuteRequest,
    current_user: UserInDB = Depends(get_current_user)
):
    """执行抽凭"""
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        # 获取规则
        cursor.execute(
            "SELECT id, name, rule_type, rule_config FROM sampling_rules WHERE id = ? AND project_id = ?",
            [request.rule_id, project_id]
        )
        rule_row = cursor.fetchone()
        if not rule_row:
            raise HTTPException(status_code=404, detail="规则不存在")

        rule_id, rule_name, rule_type, rule_config_str = rule_row
        rule_config = json.loads(rule_config_str) if isinstance(rule_config_str, str) else rule_config_str

        # 获取凭证总数
        cursor.execute("SELECT COUNT(*) FROM vouchers WHERE project_id = ?", [project_id])
        total_population = cursor.fetchone()[0]

        if total_population == 0:
            raise HTTPException(status_code=400, detail="项目中没有凭证数据")

        # 根据规则类型执行抽样
        if rule_type == "random":
            # 随机抽样
            sample_pct = rule_config.get("percentage", 10)
            sample_size = request.sample_size or max(1, int(total_population * sample_pct / 100))

            cursor.execute(
                f"""
                SELECT id, voucher_no, voucher_date, amount, subject_name, description, counterparty
                FROM vouchers
                WHERE project_id = ?
                ORDER BY RANDOM()
                LIMIT ?
                """,
                [project_id, sample_size]
            )

        elif rule_type == "amount":
            # 金额抽样
            min_amount = rule_config.get("min_amount", 0)
            max_amount = rule_config.get("max_amount")
            sample_pct = rule_config.get("percentage", 10)

            query = """
                SELECT id, voucher_no, voucher_date, amount, subject_name, description, counterparty
                FROM vouchers
                WHERE project_id = ? AND amount >= ?
            """
            params = [project_id, min_amount]

            if max_amount:
                query += " AND amount <= ?"
                params.append(max_amount)

            query += " ORDER BY amount DESC"

            cursor.execute(query, params)
            all_matches = cursor.fetchall()

            # 计算样本量
            sample_size = request.sample_size or max(1, int(len(all_matches) * sample_pct / 100))
            selected = all_matches[:sample_size]

        elif rule_type == "subject":
            # 科目抽样
            subject_codes = rule_config.get("subject_codes", [])
            if not subject_codes:
                raise HTTPException(status_code=400, detail="科目抽样规则需要指定科目代码")

            placeholders = ",".join(["?" for _ in subject_codes])
            cursor.execute(
                f"""
                SELECT id, voucher_no, voucher_date, amount, subject_name, description, counterparty
                FROM vouchers
                WHERE project_id = ? AND subject_code IN ({placeholders})
                """,
                [project_id] + subject_codes
            )
            all_matches = cursor.fetchall()
            sample_size = len(all_matches)
            selected = all_matches

        elif rule_type == "date":
            # 日期范围抽样
            start_date = rule_config.get("start_date")
            end_date = rule_config.get("end_date")

            if not start_date or not end_date:
                raise HTTPException(status_code=400, detail="日期抽样规则需要指定开始和结束日期")

            cursor.execute(
                """
                SELECT id, voucher_no, voucher_date, amount, subject_name, description, counterparty
                FROM vouchers
                WHERE project_id = ? AND voucher_date BETWEEN ? AND ?
                """,
                [project_id, start_date, end_date]
            )
            all_matches = cursor.fetchall()
            sample_size = len(all_matches)
            selected = all_matches

        elif rule_type == "stratified":
            # 分层抽样
            stratify_by = rule_config.get("stratify_by", "amount")
            layers_config = rule_config.get("layers", [])

            logger.info(f"[分层抽样] 分层配置: {layers_config}")

            if not layers_config:
                raise HTTPException(status_code=400, detail="分层抽样规则需要指定分层配置")

            # 获取所有凭证
            cursor.execute(
                """
                SELECT id, voucher_no, voucher_date, amount, subject_name, description, counterparty
                FROM vouchers
                WHERE project_id = ?
                """,
                [project_id]
            )
            all_vouchers = cursor.fetchall()
            logger.info(f"[分层抽样] 获取到 {len(all_vouchers)} 张凭证")

            # 打印凭证金额分布
            if all_vouchers:
                amounts = [float(v[3]) if v[3] else 0 for v in all_vouchers]
                logger.info(f"[分层抽样] 凭证金额分布: min={min(amounts):.2f}, max={max(amounts):.2f}, avg={sum(amounts)/len(amounts):.2f}")

            # 按层分组并抽样
            selected = []
            sample_size = 0

            # 先分配凭证到层，确保每个凭证只属于一个层
            # 按金额降序排序，优先将凭证分配到高金额层
            # 注意：SQL返回的列顺序是 id, voucher_no, voucher_date, amount, subject_name, description, counterparty
            # 所以金额在索引3，不是索引2
            sorted_vouchers = sorted(all_vouchers, key=lambda v: float(v[3]) if v[3] else 0, reverse=True)

            # 初始化各层凭证列表
            layer_vouchers_lists = [[] for _ in layers_config]
            voucher_assigned = {v[0]: False for v in all_vouchers}  # 标记凭证是否已分配

            for v in sorted_vouchers:
                voucher_id = v[0]
                voucher_no = v[1]
                amount = float(v[3]) if v[3] else 0  # 金额在索引3

                # 检查凭证属于哪些层
                matching_layers = []
                for layer_idx, layer in enumerate(layers_config):
                    layer_name = layer.get("name", "未命名层")
                    # 兼容驼峰命名和下划线命名，正确处理0值
                    min_amount = layer.get("min_amount") if layer.get("min_amount") is not None else layer.get("minAmount")
                    max_amount = layer.get("max_amount") if layer.get("max_amount") is not None else layer.get("maxAmount")

                    in_layer = True
                    if min_amount is not None and amount < min_amount:
                        in_layer = False
                    if max_amount is not None and amount >= max_amount:
                        in_layer = False

                    logger.debug(f"[分层抽样] 凭证{voucher_no}金额{amount:.2f} vs 层'{layer_name}' [{min_amount}, {max_amount}): in_layer={in_layer}")

                    if in_layer:
                        matching_layers.append(layer_idx)

                if matching_layers:
                    # 如果有多个匹配的层，选择第一个（按配置顺序）
                    # 由于凭证已按金额降序排序，高金额凭证会先被分配到高金额层
                    assigned_layer = matching_layers[0]
                    layer_vouchers_lists[assigned_layer].append(v)
                    voucher_assigned[voucher_id] = True

            # 统计未分配的凭证（通常不会有，除非配置覆盖不全）
            unassigned_count = sum(1 for assigned in voucher_assigned.values() if not assigned)
            if unassigned_count > 0:
                # 找出未分配的凭证
                unassigned_vouchers = [v for v in sorted_vouchers if not voucher_assigned[v[0]]]
                unassigned_amounts = [(v[1], float(v[3]) if v[3] else 0) for v in unassigned_vouchers]
                logger.warning(f"[分层抽样] ⚠️ {unassigned_count} 张凭证未被分配到任何层")
                logger.warning(f"[分层抽样] 未分配凭证: {unassigned_amounts}")
            else:
                logger.info(f"[分层抽样] ✅ 所有凭证已分配到各层")

            # 打印各层分配结果
            for layer_idx, layer in enumerate(layers_config):
                layer_name = layer.get("name", "未命名层")
                layer_count = len(layer_vouchers_lists[layer_idx])
                logger.info(f"[分层抽样] 层'{layer_name}': {layer_count} 张凭证")

            # 对各层进行抽样
            for layer_idx, layer in enumerate(layers_config):
                layer_name = layer.get("name", "未命名层")
                rate = layer.get("sampling_rate", 0.1)
                layer_vouchers = layer_vouchers_lists[layer_idx]

                if not layer_vouchers:
                    continue

                # 随机抽取
                import random
                random.shuffle(layer_vouchers)

                # 计算样本数
                if rate >= 1:
                    layer_sample_count = len(layer_vouchers)
                else:
                    raw_sample = len(layer_vouchers) * rate
                    # 四舍五入
                    if raw_sample < 0.5:
                        layer_sample_count = 0
                    else:
                        layer_sample_count = max(1, round(raw_sample))

                    # 确保不超过层内凭证数
                    layer_sample_count = min(layer_sample_count, len(layer_vouchers))

                if layer_sample_count > 0:
                    layer_samples = layer_vouchers[:layer_sample_count]
                    selected.extend(layer_samples)
                    sample_size += len(layer_samples)

        else:
            raise HTTPException(status_code=400, detail=f"不支持的规则类型: {rule_type}")

        # 如果是random类型，结果在cursor中
        if rule_type == "random":
            selected = cursor.fetchall()

        # 创建抽样记录ID（先创建记录用于外键引用）
        record_id = str(uuid.uuid4())
        now = datetime.now()

        # 先插入抽样记录（用于外键约束）
        cursor.execute(
            """
            INSERT INTO sampling_records (id, project_id, rule_id, rule_name, rule_type,
                                         sample_size, high_risk_count, medium_risk_count, low_risk_count, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [record_id, project_id, rule_id, rule_name, rule_type, len(selected), 0, 0, 0, 'completed', now]
        )

        # 提取凭证ID用于AI分析
        voucher_ids = [row[0] for row in selected]
        logger.info(f"[执行抽样] 抽取到 {len(voucher_ids)} 张凭证，准备调用AI分析")
        logger.info(f"[执行抽样] 凭证ID列表: {voucher_ids[:5]}{'...' if len(voucher_ids) > 5 else ''}")

        # 调用AI风险分析
        ai_results = {}
        try:
            if voucher_ids:
                logger.info(f"[执行抽样] 开始调用 analyze_vouchers_with_ai...")
                ai_results = await analyze_vouchers_with_ai(voucher_ids, project_id)
                logger.info(f"[执行抽样] AI分析返回 {len(ai_results)} 条结果")
                if ai_results:
                    sample_keys = list(ai_results.keys())[:3]
                    logger.info(f"[执行抽样] 返回的voucher_id示例: {sample_keys}")
            else:
                logger.warning(f"[执行抽样] 没有凭证需要分析")
        except Exception as e:
            logger.error(f"[执行抽样] AI风险分析失败: {str(e)}")
            import traceback
            logger.error(f"[执行抽样] 堆栈: {traceback.format_exc()}")

        # 统计风险计数
        high_risk_count = 0
        medium_risk_count = 0
        low_risk_count = 0

        # 保存抽样结果
        samples = []

        for row in selected:
            sample_id = str(uuid.uuid4())
            voucher_id = row[0]
            amount = float(row[3]) if row[3] else 0

            # 优先使用AI分析结果，否则使用基于金额的降级方案
            if voucher_id in ai_results:
                logger.info(f"[执行抽样] ✅ 使用AI结果: {voucher_id}")
                ai_result = ai_results[voucher_id]
                risk_score = ai_result.get("risk_score")
                risk_level = ai_result.get("risk_level")
                risk_factors = ai_result.get("risk_factors", [])
                explanation = ai_result.get("explanation", "")

                # 如果AI没有返回风险分数，使用金额计算
                if risk_score is None:
                    risk_score = 30  # 基础分
                    if amount > 100000:
                        risk_score = 80
                    elif amount > 50000:
                        risk_score = 60
                    elif amount > 10000:
                        risk_score = 40
            else:
                logger.info(f"[执行抽样] ⚠️ 使用降级方案: {voucher_id} (不在AI结果中)")
                # 基于金额计算风险分数
                risk_score = 30  # 基础分
                if amount > 100000:
                    risk_score = 80
                elif amount > 50000:
                    risk_score = 60
                elif amount > 10000:
                    risk_score = 40

                # 使用降级方案计算风险详情
                voucher_data = {
                    'voucher_no': row[1],
                    'voucher_date': str(row[2]) if row[2] else None,
                    'amount': amount,
                    'subject_name': row[4],
                    'description': row[5],
                    'counterparty': row[6] if len(row) > 6 else None
                }
                risk_details = calculate_risk_details(risk_score, amount, voucher_data)
                risk_level = risk_details["risk_level"]
                risk_factors = risk_details["risk_factors"]
                explanation = risk_details["explanation"]

            # 统计风险计数
            if risk_level == "high":
                high_risk_count += 1
            elif risk_level == "medium":
                medium_risk_count += 1
            else:
                low_risk_count += 1

            # 插入样本记录
            cursor.execute(
                """
                INSERT INTO samples (id, project_id, record_id, rule_id, voucher_id, risk_score, risk_level, risk_factors, explanation, reason, sampled_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [sample_id, project_id, record_id, rule_id, voucher_id, risk_score, risk_level,
                 json.dumps(risk_factors, ensure_ascii=False) if risk_factors else None,
                 explanation, f"规则【{rule_name}】抽取", now]
            )

            # 构建样本响应
            samples.append(SampleResponse(
                id=sample_id,
                voucher_id=voucher_id,
                voucher_no=row[1],
                amount=float(row[3]) if row[3] else None,
                subject_name=row[4],
                description=row[5],
                risk_score=risk_score,
                risk_level=risk_level,
                risk_factors=risk_factors,
                explanation=explanation,
                reason=f"规则【{rule_name}】抽取"
            ))

        # 更新抽样记录的风险统计
        cursor.execute(
            """
            UPDATE sampling_records
            SET high_risk_count = ?, medium_risk_count = ?, low_risk_count = ?
            WHERE id = ?
            """,
            [high_risk_count, medium_risk_count, low_risk_count, record_id]
        )

        get_db().commit()

        return SamplingResultResponse(
            rule_id=rule_id,
            rule_name=rule_name,
            total_population=total_population,
            sample_size=len(samples),
            samples=samples
        )


@router.get("/projects/{project_id}/samples")
async def list_samples(
    project_id: str,
    current_user: UserInDB = Depends(get_current_user),
    rule_id: Optional[str] = Query(None, description="按规则筛选"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200)
):
    """获取抽凭结果"""
    offset = (page - 1) * page_size

    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        # 构建查询
        conditions = ["s.project_id = ?"]
        params = [project_id]

        if rule_id:
            conditions.append("s.rule_id = ?")
            params.append(rule_id)

        where_clause = " AND ".join(conditions)

        # 查询总数
        cursor.execute(
            f"""
            SELECT COUNT(*)
            FROM samples s
            WHERE {where_clause}
            """,
            params
        )
        total = cursor.fetchone()[0]

        # 查询列表
        cursor.execute(
            f"""
            SELECT s.id, s.voucher_id, v.voucher_no, v.amount, v.subject_name,
                   v.description, s.risk_score, s.reason
            FROM samples s
            JOIN vouchers v ON s.voucher_id = v.id
            WHERE {where_clause}
            ORDER BY s.sampled_at DESC
            LIMIT ? OFFSET ?
            """,
            params + [page_size, offset]
        )
        rows = cursor.fetchall()

        items = [
            SampleResponse(
                id=row[0],
                voucher_id=row[1],
                voucher_no=row[2],
                amount=float(row[3]) if row[3] else None,
                subject_name=row[4],
                description=row[5],
                risk_score=float(row[6]) if row[6] else None,
                reason=row[7]
            )
            for row in rows
        ]

        return {"total": total, "items": items}


@router.post("/projects/{project_id}/samples/export")
async def export_samples(
    project_id: str,
    current_user: UserInDB = Depends(get_current_user),
    rule_id: Optional[str] = Query(None, description="规则ID"),
    format: str = Query("excel", description="导出格式: excel/pdf")
):
    """
    导出抽凭结果

    Args:
        project_id: 项目ID
        rule_id: 规则ID（可选）
        format: 导出格式 (excel/pdf)
    """
    from fastapi.responses import Response
    from app.services.export_service import excel_exporter, pdf_exporter

    # 验证项目存在
    with get_db_cursor() as cursor:
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

    try:
        if format == "pdf":
            file_data = pdf_exporter.export_working_paper(project_id)
            filename = f"sampling_report_{project_id}.pdf"
            media_type = "application/pdf"
        else:
            file_data = excel_exporter.export_sampling_results(project_id, rule_id)
            filename = f"sampling_results_{project_id}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        return Response(
            content=file_data,
            media_type=media_type,
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    except RuntimeError as e:
        logger.error(f"导出失败: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        logger.error(f"导出失败: {str(e)}")
        raise HTTPException(status_code=500, detail="导出失败，请稍后重试")


@router.post("/projects/{project_id}/sampling/stratified-preview")
async def preview_stratified_sampling(
    project_id: str,
    config: StratifiedSamplingConfig,
    current_user: UserInDB = Depends(get_current_user)
):
    """
    预览分层抽样结果（不实际执行抽样，只返回预估统计）

    Args:
        project_id: 项目ID
        config: 分层抽样配置
    """
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        # 获取所有凭证
        cursor.execute(
            """
            SELECT id, amount
            FROM vouchers
            WHERE project_id = ?
            """,
            [project_id]
        )
        all_vouchers = cursor.fetchall()

        total_population = len(all_vouchers)

        if total_population == 0:
            raise HTTPException(status_code=400, detail="项目中没有凭证数据")

        # 计算各层统计
        layers_summary = []
        total_sample_size = 0

        # 记录每张凭证所属的层，用于验证互斥性
        voucher_layers = {v[0]: [] for v in all_vouchers}  # 凭证ID -> 层索引列表

        # 第一遍：分配凭证到各层，并检查重叠
        layer_counts = []
        for layer_idx, layer in enumerate(config.layers):
            layer_name = layer.name
            min_amount = layer.min_amount
            max_amount = layer.max_amount
            rate = layer.sampling_rate

            layer_count = 0
            for v in all_vouchers:
                voucher_id = v[0]
                amount = float(v[1]) if v[1] else 0
                in_layer = True

                if min_amount is not None and amount < min_amount:
                    in_layer = False
                if max_amount is not None and amount >= max_amount:
                    in_layer = False

                if in_layer:
                    layer_count += 1
                    voucher_layers[voucher_id].append(layer_idx)

            layer_counts.append((layer_name, layer_count, min_amount, max_amount, rate))

        # 检查重叠：统计属于多层的凭证
        overlapping_vouchers = 0
        for voucher_id, layers in voucher_layers.items():
            if len(layers) > 1:
                overlapping_vouchers += 1

        if overlapping_vouchers > 0:
            # 记录警告，但继续执行（可能是配置问题）
            import logging
            logging.warning(f"分层配置存在重叠: {overlapping_vouchers} 张凭证属于多个层")

        # 第二遍：计算样本量
        for layer_idx, (layer_name, layer_count, min_amount, max_amount, rate) in enumerate(layer_counts):
            # 计算样本数
            if layer_count == 0:
                sample_count = 0
            elif rate >= 1:
                sample_count = layer_count
            else:
                # 使用四舍五入而不是向下取整
                raw_sample = layer_count * rate
                # 如果抽样结果小于0.5，且层内凭证数较少，可能不抽样
                # 但审计中通常至少抽1个
                if raw_sample < 0.5:
                    sample_count = 0
                else:
                    sample_count = max(1, round(raw_sample))

                # 确保样本数不超过层内凭证数
                sample_count = min(sample_count, layer_count)

            layers_summary.append({
                "name": layer_name,
                "population_count": layer_count,
                "sample_count": sample_count,
                "sampling_rate": rate
            })

            total_sample_size += sample_count

        # 确保总样本数不超过总体
        if total_sample_size > total_population:
            # 按比例调整各层样本数
            scale = total_population / total_sample_size
            adjusted_total = 0
            for layer in layers_summary:
                if layer["sample_count"] > 0:
                    adjusted = max(1, round(layer["sample_count"] * scale))
                    adjusted = min(adjusted, layer["population_count"])
                    layer["sample_count"] = adjusted
                adjusted_total += layer["sample_count"]

            # 递归调整直到不超过总体（通常一次就够）
            while adjusted_total > total_population:
                for layer in layers_summary:
                    if layer["sample_count"] > 1:
                        layer["sample_count"] -= 1
                        adjusted_total -= 1
                        if adjusted_total <= total_population:
                            break

            total_sample_size = adjusted_total

        return {
            "total_population": total_population,
            "total_sample_size": total_sample_size,
            "overall_sampling_rate": round(total_sample_size / total_population * 100, 2) if total_population > 0 else 0,
            "layers_summary": layers_summary,
            "warning": f"发现 {overlapping_vouchers} 张凭证属于多个层" if overlapping_vouchers > 0 else None
        }


# ==================== 抽样记录API ====================

class SamplingRecordResponse(BaseModel):
    """抽样记录响应"""
    id: str
    project_id: str
    rule_id: Optional[str]
    rule_name: str
    rule_type: str
    sample_size: int
    high_risk_count: int
    medium_risk_count: int
    low_risk_count: int
    status: str
    created_at: datetime


class SamplingRecordDetail(BaseModel):
    """抽样记录详情"""
    record: SamplingRecordResponse
    samples: List[SampleResponse]


@router.get("/projects/{project_id}/sampling-records", response_model=List[SamplingRecordResponse])
async def list_sampling_records(
    project_id: str,
    current_user: UserInDB = Depends(get_current_user)
):
    """获取抽样记录列表"""
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        cursor.execute(
            """
            SELECT id, project_id, rule_id, rule_name, rule_type, sample_size,
                   high_risk_count, medium_risk_count, low_risk_count, status, created_at
            FROM sampling_records
            WHERE project_id = ?
            ORDER BY created_at DESC
            """,
            [project_id]
        )
        rows = cursor.fetchall()

        return [
            SamplingRecordResponse(
                id=row[0],
                project_id=row[1],
                rule_id=row[2],
                rule_name=row[3],
                rule_type=row[4],
                sample_size=row[5] or 0,
                high_risk_count=row[6] or 0,
                medium_risk_count=row[7] or 0,
                low_risk_count=row[8] or 0,
                status=row[9] or 'completed',
                created_at=row[10]
            )
            for row in rows
        ]


@router.get("/projects/{project_id}/sampling-records/{record_id}", response_model=SamplingRecordDetail)
async def get_sampling_record_detail(
    project_id: str,
    record_id: str,
    current_user: UserInDB = Depends(get_current_user)
):
    """获取抽样记录详情"""
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        # 获取记录
        cursor.execute(
            """
            SELECT id, project_id, rule_id, rule_name, rule_type, sample_size,
                   high_risk_count, medium_risk_count, low_risk_count, status, created_at
            FROM sampling_records
            WHERE id = ? AND project_id = ?
            """,
            [record_id, project_id]
        )
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="抽样记录不存在")

        record = SamplingRecordResponse(
            id=row[0],
            project_id=row[1],
            rule_id=row[2],
            rule_name=row[3],
            rule_type=row[4],
            sample_size=row[5] or 0,
            high_risk_count=row[6] or 0,
            medium_risk_count=row[7] or 0,
            low_risk_count=row[8] or 0,
            status=row[9] or 'completed',
            created_at=row[10]
        )

        # 获取样本明细，从vouchers表获取完整的AI分析结果
        try:
            cursor.execute(
                """
                SELECT s.id, s.voucher_id, v.voucher_no, v.voucher_date, v.amount, v.subject_name,
                       v.description, v.counterparty, s.risk_score, s.risk_level,
                       s.risk_factors, s.explanation, s.reason,
                       v.risk_tags, v.ai_analysis
                FROM samples s
                JOIN vouchers v ON s.voucher_id = v.id
                WHERE s.record_id = ?
                ORDER BY s.risk_score DESC NULLS LAST, v.amount DESC
                """,
                [record_id]
            )
            has_voucher_risk = True
        except Exception:
            # 如果列不存在，使用基本查询
            cursor.execute(
                """
                SELECT s.id, s.voucher_id, v.voucher_no, v.amount, v.subject_name,
                       v.description, s.risk_score, s.reason
                FROM samples s
                JOIN vouchers v ON s.voucher_id = v.id
                WHERE s.record_id = ?
                ORDER BY s.risk_score DESC NULLS LAST, v.amount DESC
                """,
                [record_id]
            )
            has_voucher_risk = False
        sample_rows = cursor.fetchall()

        samples = []
        for row in sample_rows:
            if has_voucher_risk:
                # 从vouchers表获取完整的AI分析结果
                sample_id = row[0]
                voucher_id = row[1]
                voucher_no = row[2]
                voucher_date = str(row[3]) if row[3] else None
                amount = float(row[4]) if row[4] else None
                subject_name = row[5]
                description = row[6]
                counterparty = row[7]
                risk_score = float(row[8]) if row[8] else None
                risk_level = row[9] or "low"
                sample_risk_factors = json.loads(row[10]) if row[10] else []
                sample_explanation = row[11]
                reason = row[12]
                voucher_risk_tags = json.loads(row[13]) if row[13] else []
                voucher_ai_analysis = json.loads(row[14]) if row[14] else {}

                # 构建风险因素：只从risk_tags获取，不包含AI分析和AI建议
                risk_factors = sample_risk_factors.copy() if sample_risk_factors else []

                # 从risk_tags提取风险因素名称
                if voucher_risk_tags and not risk_factors:
                    for tag in voucher_risk_tags:
                        if isinstance(tag, dict):
                            tag_name = tag.get('name', '')
                            if tag_name:
                                risk_factors.append(tag_name)
                        elif isinstance(tag, str):
                            risk_factors.append(tag)

                # 构建explanation：包含AI分析和AI建议
                explanation_parts = []
                if voucher_ai_analysis:
                    ai_explanation = voucher_ai_analysis.get('explanation', '')
                    ai_attention = voucher_ai_analysis.get('audit_attention', [])

                    # AI分析说明
                    if ai_explanation:
                        explanation_parts.append(f"AI分析: {ai_explanation}")

                    # AI建议
                    if ai_attention:
                        for attention in ai_attention:
                            if attention:
                                explanation_parts.append(f"AI建议: {attention}")

                # 设置explanation：优先使用AI分析结果，否则降级
                if sample_explanation:
                    explanation = sample_explanation
                elif explanation_parts:
                    explanation = "\n".join(explanation_parts)
                else:
                    # 最后降级使用规则计算
                    risk_details = calculate_risk_details(risk_score, amount)
                    explanation = risk_details["explanation"]
                    if not risk_factors:
                        risk_factors = risk_details["risk_factors"]
            else:
                # 无完整风险详情字段，使用计算的值
                risk_score = float(row[6]) if row[6] else None
                amount = float(row[3]) if row[3] else None
                risk_details = calculate_risk_details(risk_score, amount)
                risk_level = risk_details["risk_level"]
                risk_factors = risk_details["risk_factors"]
                explanation = risk_details["explanation"]
                reason = row[7] if len(row) > 7 else None

            samples.append(SampleResponse(
                id=sample_id if has_voucher_risk else row[0],
                voucher_id=voucher_id if has_voucher_risk else row[1],
                voucher_no=voucher_no if has_voucher_risk else row[2],
                amount=amount,
                subject_name=subject_name if has_voucher_risk else row[4],
                description=description if has_voucher_risk else row[5],
                risk_score=risk_score,
                risk_level=risk_level,
                risk_factors=risk_factors,
                explanation=explanation,
                reason=reason
            ))

        return SamplingRecordDetail(record=record, samples=samples)


@router.delete("/projects/{project_id}/sampling-records/{record_id}")
async def delete_sampling_record(
    project_id: str,
    record_id: str,
    current_user: UserInDB = Depends(get_current_user)
):
    """删除抽样记录"""
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        # 验证记录存在
        cursor.execute(
            "SELECT id, rule_id FROM sampling_records WHERE id = ? AND project_id = ?",
            [record_id, project_id]
        )
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="抽样记录不存在")

        rule_id = row[1]

        # 删除样本
        cursor.execute("DELETE FROM samples WHERE record_id = ?", [record_id])

        # 删除抽样记录
        cursor.execute("DELETE FROM sampling_records WHERE id = ?", [record_id])

        # 删除规则（如果没有其他记录使用）
        cursor.execute(
            "SELECT COUNT(*) FROM sampling_records WHERE rule_id = ?",
            [rule_id]
        )
        if cursor.fetchone()[0] == 0:
            cursor.execute("DELETE FROM sampling_rules WHERE id = ?", [rule_id])

        get_db().commit()

    return {"message": "抽样记录已删除", "record_id": record_id}


@router.post("/projects/{project_id}/sampling-records/{record_id}/export")
async def export_sampling_record(
    project_id: str,
    record_id: str,
    current_user: UserInDB = Depends(get_current_user),
    format: str = Query("excel", description="导出格式: excel/pdf")
):
    """
    导出抽样记录详情
    """
    from fastapi.responses import Response
    from io import BytesIO

    with get_db_cursor() as cursor:
        # 验证项目和记录存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        cursor.execute(
            """
            SELECT id, rule_name, sample_size, high_risk_count, medium_risk_count, low_risk_count, created_at
            FROM sampling_records
            WHERE id = ? AND project_id = ?
            """,
            [record_id, project_id]
        )
        record_row = cursor.fetchone()
        if not record_row:
            raise HTTPException(status_code=404, detail="抽样记录不存在")

        # 获取样本明细
        cursor.execute(
            """
            SELECT s.id, v.voucher_no, v.voucher_date, v.amount, v.subject_name,
                   v.description, s.risk_score, s.reason
            FROM samples s
            JOIN vouchers v ON s.voucher_id = v.id
            WHERE s.record_id = ?
            ORDER BY s.risk_score DESC NULLS LAST, v.amount DESC
            """,
            [record_id]
        )
        sample_rows = cursor.fetchall()

    try:
        if format == "pdf":
            # PDF导出（简化版）
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            # 尝试注册中文字体
            try:
                pdfmetrics.registerFont(TTFont('SimHei', 'simhei.ttf'))
                font_name = 'SimHei'
            except:
                font_name = 'Helvetica'

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
            elements = []

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=18
            )

            # 标题
            elements.append(Paragraph(f"抽样结果报告 - {record_row[1]}", title_style))
            elements.append(Spacer(1, 20))

            # 统计信息
            stats_data = [
                ['样本总数', '高风险', '中风险', '低风险'],
                [str(record_row[2]), str(record_row[3]), str(record_row[4]), str(record_row[5])]
            ]
            stats_table = Table(stats_data, colWidths=[100, 100, 100, 100])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ]))
            elements.append(stats_table)
            elements.append(Spacer(1, 20))

            # 样本明细表
            data = [['序号', '凭证号', '日期', '金额', '科目', '摘要', '风险分数', '抽取原因']]
            for i, row in enumerate(sample_rows, 1):
                data.append([
                    str(i),
                    str(row[1] or ''),
                    str(row[2] or ''),
                    f"{float(row[3]):,.2f}" if row[3] else '0.00',
                    str(row[4] or ''),
                    str(row[5] or '')[:30],
                    f"{float(row[6]):.1f}" if row[6] else '-',
                    str(row[7] or '')
                ])

            table = Table(data, colWidths=[40, 80, 80, 100, 100, 150, 80, 120])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)

            doc.build(elements)
            buffer.seek(0)

            return Response(
                content=buffer.getvalue(),
                media_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="sampling_record_{record_id}.pdf"'}
            )
        else:
            # Excel导出
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "抽样结果"

            # 标题
            ws.merge_cells('A1:H1')
            ws['A1'] = f"抽样结果报告 - {record_row[1]}"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')

            # 统计信息
            ws['A3'] = '样本总数'
            ws['B3'] = record_row[2]
            ws['C3'] = '高风险'
            ws['D3'] = record_row[3]
            ws['E3'] = '中风险'
            ws['F3'] = record_row[4]
            ws['G3'] = '低风险'
            ws['H3'] = record_row[5]

            # 表头
            headers = ['序号', '凭证号', '日期', '金额', '科目', '摘要', '风险分数', '抽取原因']
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # 数据
            for i, row in enumerate(sample_rows, 1):
                ws.cell(row=5+i, column=1, value=i)
                ws.cell(row=5+i, column=2, value=str(row[1] or ''))
                ws.cell(row=5+i, column=3, value=str(row[2] or ''))
                ws.cell(row=5+i, column=4, value=float(row[3]) if row[3] else 0)
                ws.cell(row=5+i, column=5, value=str(row[4] or ''))
                ws.cell(row=5+i, column=6, value=str(row[5] or ''))
                ws.cell(row=5+i, column=7, value=float(row[6]) if row[6] else None)
                ws.cell(row=5+i, column=8, value=str(row[7] or ''))

            # 调整列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 20

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return Response(
                content=buffer.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="sampling_record_{record_id}.xlsx"'}
            )

    except Exception as e:
        logger.error(f"导出失败: {str(e)}")
        raise HTTPException(status_code=500, detail="导出失败，请稍后重试")


# ==================== MUS抽样API ====================

class MUSPreviewRequest(BaseModel):
    """MUS预览请求"""
    confidence_level: float = Field(0.95, ge=0.90, le=0.99, description="置信水平")
    tolerable_misstatement: float = Field(0.05, ge=0.01, le=0.20, description="可容忍错报率")
    expected_misstatement: float = Field(0.01, ge=0, le=0.10, description="预期错报率")


class MUSPreviewResponse(BaseModel):
    """MUS预览响应"""
    sample_size: int
    sampling_interval: float
    random_start: float
    reliability_factor: float
    tolerable_amount: float
    confidence_level: float
    population_amount: float


@router.post("/projects/{project_id}/sampling/mus-preview", response_model=MUSPreviewResponse)
async def preview_mus_sampling(
    project_id: str,
    config: MUSPreviewRequest,
    current_user: UserInDB = Depends(get_current_user)
):
    """
    预览MUS抽样参数

    计算货币单位抽样所需的样本量、抽样间隔等参数
    """
    from app.services.mus_sampling_service import MUSSampler

    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        # 获取总体金额
        cursor.execute(
            "SELECT COALESCE(SUM(amount), 0), COUNT(*) FROM vouchers WHERE project_id = ?",
            [project_id]
        )
        row = cursor.fetchone()
        population_amount = float(row[0]) if row[0] else 0
        population_count = int(row[1]) if row[1] else 0

    if population_amount <= 0:
        raise HTTPException(status_code=400, detail="项目中没有凭证金额数据，无法进行MUS抽样")

    sampler = MUSSampler()
    result = sampler.calculate_sample_size(
        population_amount=population_amount,
        confidence_level=config.confidence_level,
        tolerable_misstatement=config.tolerable_misstatement,
        expected_misstatement=config.expected_misstatement
    )

    # MUS样本量是货币单位选择点数，实际样本数不超过总体凭证数
    # 一张凭证可能被多次选中，但实际抽取的是凭证本身
    theoretical_sample_size = result["sample_size"]
    actual_sample_size = min(theoretical_sample_size, population_count)

    # 计算实际抽样率
    sampling_rate = round((actual_sample_size / population_count * 100), 1) if population_count > 0 else 0

    return MUSPreviewResponse(
        sample_size=actual_sample_size,
        sampling_interval=result["sampling_interval"],
        random_start=result["random_start"],
        reliability_factor=result["reliability_factor"],
        tolerable_amount=result["tolerable_amount"],
        confidence_level=result["confidence_level"],
        population_amount=result["population_amount"]
    )


class MUSExecuteRequest(BaseModel):
    """MUS执行请求"""
    name: str = Field(..., description="规则名称")
    confidence_level: float = Field(0.95, ge=0.90, le=0.99)
    tolerable_misstatement: float = Field(0.05, ge=0.01, le=0.20)
    expected_misstatement: float = Field(0.01, ge=0, le=0.10)
    sampling_interval: Optional[float] = Field(None, description="抽样间隔（可选，系统自动计算）")
    random_start: Optional[float] = Field(None, description="随机起点（可选，系统自动生成）")


@router.post("/projects/{project_id}/sampling/mus-execute")
async def execute_mus_sampling(
    project_id: str,
    request: MUSExecuteRequest,
    current_user: UserInDB = Depends(get_current_user)
):
    """
    执行MUS抽样

    按货币单位抽样方法抽取样本
    """
    from app.services.mus_sampling_service import MUSSampler

    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        # 获取总体信息
        cursor.execute(
            "SELECT COALESCE(SUM(amount), 0), COUNT(*) FROM vouchers WHERE project_id = ?",
            [project_id]
        )
        row = cursor.fetchone()
        population_amount = float(row[0]) if row[0] else 0
        population_count = int(row[1]) if row[1] else 0

        if population_amount <= 0:
            raise HTTPException(status_code=400, detail="项目中没有凭证金额数据")

        # 计算抽样参数
        sampler = MUSSampler()
        params = sampler.calculate_sample_size(
            population_amount=population_amount,
            confidence_level=request.confidence_level,
            tolerable_misstatement=request.tolerable_misstatement,
            expected_misstatement=request.expected_misstatement
        )

        # 使用传入的参数或计算的参数
        interval = request.sampling_interval or params["sampling_interval"]
        random_start = request.random_start if request.random_start is not None else params["random_start"]

        # 获取所有凭证
        cursor.execute(
            """
            SELECT id, voucher_no, amount, subject_name, description, counterparty
            FROM vouchers
            WHERE project_id = ? AND amount > 0
            ORDER BY amount DESC
            """,
            [project_id]
        )
        all_vouchers = [
            {
                "id": row[0],
                "voucher_no": row[1],
                "amount": float(row[2]) if row[2] else 0,
                "subject_name": row[3],
                "description": row[4]
            }
            for row in cursor.fetchall()
        ]

        # 执行MUS抽样
        selected = sampler.select_samples(all_vouchers, interval, random_start)

        # MUS抽样去重：同一张凭证只保留一条记录，但记录被选中次数
        seen_vouchers = {}
        for item in selected:
            vid = item["id"]
            if vid not in seen_vouchers:
                seen_vouchers[vid] = item
                seen_vouchers[vid]["selection_count"] = 1
                seen_vouchers[vid]["selection_points"] = [item.get("selection_point", 0)]
            else:
                seen_vouchers[vid]["selection_count"] += 1
                seen_vouchers[vid]["selection_points"].append(item.get("selection_point", 0))

        # 去重后的样本列表
        unique_selected = list(seen_vouchers.values())

        # 创建规则
        rule_id = str(uuid.uuid4())
        rule_config = {
            "confidence_level": request.confidence_level,
            "tolerable_misstatement": request.tolerable_misstatement,
            "expected_misstatement": request.expected_misstatement,
            "sampling_interval": interval,
            "random_start": random_start,
            "reliability_factor": params["reliability_factor"]
        }

        cursor.execute(
            """
            INSERT INTO sampling_rules (id, project_id, name, rule_type, rule_config,
                                        confidence_level, tolerable_error, expected_error,
                                        sampling_interval, random_start, population_amount, is_active, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [rule_id, project_id, request.name, "monetary_unit", json.dumps(rule_config),
             request.confidence_level, request.tolerable_misstatement, request.expected_misstatement,
             interval, random_start, population_amount, True, datetime.now()]
        )

        # 创建抽样记录
        record_id = str(uuid.uuid4())
        now = datetime.now()

        cursor.execute(
            """
            INSERT INTO sampling_records (id, project_id, rule_id, rule_name, rule_type,
                                         sample_size, population_count, population_amount, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [record_id, project_id, rule_id, request.name, "monetary_unit",
             len(unique_selected), population_count, population_amount, 'completed', now]
        )

        # 保存样本
        samples = []
        high_risk_count = 0
        medium_risk_count = 0
        low_risk_count = 0

        for item in unique_selected:
            sample_id = str(uuid.uuid4())

            # 基于金额计算风险分数
            amount = item.get("amount", 0)
            risk_score = 30  # 基础分
            if amount > 100000:
                risk_score = 80
            elif amount > 50000:
                risk_score = 60
            elif amount > 10000:
                risk_score = 40

            # 构建原因说明（包含选中次数）
            selection_count = item.get("selection_count", 1)
            reason = f"MUS抽样 - 选中{selection_count}次"
            if selection_count > 1:
                points = item.get("selection_points", [])
                reason += f"，选择点: {', '.join([f'{p:.2f}' for p in points[:3]])}"
                if len(points) > 3:
                    reason += "..."

            # 计算风险详情
            risk_details = calculate_risk_details(risk_score, amount)

            # 统计风险计数
            if risk_details["risk_level"] == "high":
                high_risk_count += 1
            elif risk_details["risk_level"] == "medium":
                medium_risk_count += 1
            else:
                low_risk_count += 1

            cursor.execute(
                """
                INSERT INTO samples (id, project_id, record_id, rule_id, voucher_id, risk_score, risk_level, risk_factors, explanation, reason, sampled_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [sample_id, project_id, record_id, rule_id, item["id"],
                 risk_score, risk_details["risk_level"],
                 json.dumps(risk_details["risk_factors"], ensure_ascii=False) if risk_details["risk_factors"] else None,
                 risk_details["explanation"], reason, now]
            )

            risk_details = calculate_risk_details(risk_score, amount)
            samples.append(SampleResponse(
                id=sample_id,
                voucher_id=item["id"],
                voucher_no=item["voucher_no"],
                amount=item["amount"],
                subject_name=item.get("subject_name"),
                description=item.get("description"),
                risk_score=risk_score,
                risk_level=risk_details["risk_level"],
                risk_factors=risk_details["risk_factors"],
                explanation=risk_details["explanation"],
                reason=reason
            ))

        # 更新抽样记录的风险计数
        cursor.execute(
            """
            UPDATE sampling_records
            SET high_risk_count = ?, medium_risk_count = ?, low_risk_count = ?
            WHERE id = ?
            """,
            [high_risk_count, medium_risk_count, low_risk_count, record_id]
        )

        get_db().commit()

        return SamplingResultResponse(
            rule_id=rule_id,
            rule_name=request.name,
            total_population=population_count,
            sample_size=len(samples),
            samples=samples
        )


# ==================== 系统抽样API ====================

class SystematicPreviewRequest(BaseModel):
    """系统抽样预览请求"""
    sample_size: Optional[int] = Field(None, ge=1, description="样本量")
    sampling_rate: Optional[float] = Field(None, ge=0.01, le=1.0, description="抽样比例")


class SystematicPreviewResponse(BaseModel):
    """系统抽样预览响应"""
    interval: int
    random_start: int
    population_size: int
    sample_size: int
    actual_sample_size: int
    sampling_rate: float


@router.post("/projects/{project_id}/sampling/systematic-preview", response_model=SystematicPreviewResponse)
async def preview_systematic_sampling(
    project_id: str,
    config: SystematicPreviewRequest,
    current_user: UserInDB = Depends(get_current_user)
):
    """
    预览系统抽样参数

    计算系统抽样所需的抽样间隔、随机起点等参数
    """
    from app.services.systematic_sampling_service import SystematicSampler

    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        # 获取总体规模
        cursor.execute("SELECT COUNT(*) FROM vouchers WHERE project_id = ?", [project_id])
        population_size = cursor.fetchone()[0]

    if population_size <= 0:
        raise HTTPException(status_code=400, detail="项目中没有凭证数据")

    sampler = SystematicSampler()

    if config.sample_size:
        result = sampler.calculate_parameters(population_size, config.sample_size)
    elif config.sampling_rate:
        result = sampler.calculate_sample_size_from_rate(population_size, config.sampling_rate)
    else:
        # 默认10%抽样率
        result = sampler.calculate_sample_size_from_rate(population_size, 0.1)

    return SystematicPreviewResponse(
        interval=result["interval"],
        random_start=result["random_start"],
        population_size=result["population_size"],
        sample_size=result["sample_size"],
        actual_sample_size=result.get("actual_sample_size", result["sample_size"]),
        sampling_rate=result.get("sampling_rate", result["sample_size"] / population_size * 100)
    )


class SystematicExecuteRequest(BaseModel):
    """系统抽样执行请求"""
    name: str = Field(..., description="规则名称")
    sample_size: Optional[int] = Field(None, ge=1, description="样本量")
    sampling_rate: Optional[float] = Field(None, ge=0.01, le=1.0, description="抽样比例")
    interval: Optional[int] = Field(None, ge=1, description="抽样间隔")
    random_start: Optional[int] = Field(None, ge=1, description="随机起点")


@router.post("/projects/{project_id}/sampling/systematic-execute")
async def execute_systematic_sampling(
    project_id: str,
    request: SystematicExecuteRequest,
    current_user: UserInDB = Depends(get_current_user)
):
    """
    执行系统抽样

    按固定间隔从总体中抽取样本
    """
    from app.services.systematic_sampling_service import SystematicSampler

    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        # 获取总体信息
        cursor.execute(
            "SELECT COUNT(*), COALESCE(SUM(amount), 0) FROM vouchers WHERE project_id = ?",
            [project_id]
        )
        row = cursor.fetchone()
        population_size = int(row[0]) if row[0] else 0
        population_amount = float(row[1]) if row[1] else 0

        if population_size <= 0:
            raise HTTPException(status_code=400, detail="项目中没有凭证数据")

        # 计算抽样参数
        sampler = SystematicSampler()

        if request.interval and request.random_start:
            # 使用传入的参数
            interval = request.interval
            random_start = request.random_start
            sample_size = population_size // interval
        elif request.sample_size:
            result = sampler.calculate_parameters(population_size, request.sample_size)
            interval = result["interval"]
            random_start = result["random_start"]
            sample_size = request.sample_size
        elif request.sampling_rate:
            result = sampler.calculate_sample_size_from_rate(population_size, request.sampling_rate)
            interval = result["interval"]
            random_start = result["random_start"]
            sample_size = result["sample_size"]
        else:
            # 默认10%抽样率
            result = sampler.calculate_sample_size_from_rate(population_size, 0.1)
            interval = result["interval"]
            random_start = result["random_start"]
            sample_size = result["sample_size"]

        # 获取所有凭证
        cursor.execute(
            """
            SELECT id, voucher_no, amount, subject_name, description, counterparty
            FROM vouchers
            WHERE project_id = ?
            """,
            [project_id]
        )
        all_vouchers = [
            {
                "id": row[0],
                "voucher_no": row[1],
                "amount": float(row[2]) if row[2] else 0,
                "subject_name": row[3],
                "description": row[4]
            }
            for row in cursor.fetchall()
        ]

        # 执行系统抽样
        selected = sampler.select_samples(all_vouchers, interval, random_start)

        # 创建规则
        rule_id = str(uuid.uuid4())
        rule_config = {
            "sample_size": sample_size,
            "sampling_rate": request.sampling_rate or (sample_size / population_size),
            "interval": interval,
            "random_start": random_start
        }

        cursor.execute(
            """
            INSERT INTO sampling_rules (id, project_id, name, rule_type, rule_config,
                                        sampling_interval, random_start, population_amount, is_active, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [rule_id, project_id, request.name, "systematic", json.dumps(rule_config),
             float(interval), float(random_start), population_amount, True, datetime.now()]
        )

        # 创建抽样记录
        record_id = str(uuid.uuid4())
        now = datetime.now()

        cursor.execute(
            """
            INSERT INTO sampling_records (id, project_id, rule_id, rule_name, rule_type,
                                         sample_size, population_count, population_amount, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [record_id, project_id, rule_id, request.name, "systematic",
             len(selected), population_size, population_amount, 'completed', now]
        )

        # 保存样本
        samples = []
        for item in selected:
            sample_id = str(uuid.uuid4())

            # 基于金额计算风险分数
            amount = item.get("amount", 0)
            risk_score = 30  # 基础分
            if amount > 100000:
                risk_score = 80
            elif amount > 50000:
                risk_score = 60
            elif amount > 10000:
                risk_score = 40

            # 计算风险详情
            risk_details = calculate_risk_details(risk_score, amount)

            cursor.execute(
                """
                INSERT INTO samples (id, project_id, record_id, rule_id, voucher_id, risk_score, risk_level, risk_factors, explanation, reason, sampled_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [sample_id, project_id, record_id, rule_id, item["id"],
                 risk_score, risk_details["risk_level"],
                 json.dumps(risk_details["risk_factors"], ensure_ascii=False) if risk_details["risk_factors"] else None,
                 risk_details["explanation"], f"系统抽样 - 序号:{item.get('selection_sequence', 0)}", now]
            )

            risk_details = calculate_risk_details(risk_score, amount)
            samples.append(SampleResponse(
                id=sample_id,
                voucher_id=item["id"],
                voucher_no=item["voucher_no"],
                amount=item["amount"],
                subject_name=item.get("subject_name"),
                description=item.get("description"),
                risk_score=risk_score,
                risk_level=risk_details["risk_level"],
                risk_factors=risk_details["risk_factors"],
                explanation=risk_details["explanation"],
                reason=f"系统抽样 - 序号:{item.get('selection_sequence', 0)}"
            ))

        get_db().commit()

        return SamplingResultResponse(
            rule_id=rule_id,
            rule_name=request.name,
            total_population=population_size,
            sample_size=len(samples),
            samples=samples
        )


# ==================== 样本测试API ====================

class SampleTestRequest(BaseModel):
    """样本测试请求"""
    sample_ids: List[str] = Field(..., description="样本ID列表")


class SampleTestResponse(BaseModel):
    """样本测试响应"""
    total_tested: int
    need_review_count: int
    error_count: int
    results: List[Dict[str, Any]]


@router.post("/projects/{project_id}/samples/test", response_model=SampleTestResponse)
async def execute_sample_test(
    project_id: str,
    request: SampleTestRequest,
    current_user: UserInDB = Depends(get_current_user)
):
    """
    执行AI样本测试

    对选定的样本进行批量AI风险分析和证据链校验
    """
    from app.services.sample_test_service import SampleTestService

    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

    service = SampleTestService()
    result = await service.execute_ai_test(project_id, request.sample_ids)

    return SampleTestResponse(
        total_tested=result["total_tested"],
        need_review_count=result["need_review_count"],
        error_count=result.get("error_count", 0),
        results=result["results"]
    )


class MisstatementCreate(BaseModel):
    """创建错报记录请求"""
    sample_id: str = Field(..., description="样本ID")
    misstatement_type: str = Field(..., description="错报类型")
    misstatement_amount: float = Field(..., description="错报金额")
    original_amount: float = Field(..., description="原始金额")
    correct_amount: float = Field(..., description="正确金额")
    description: str = Field(..., description="错报描述")
    evidence_path: Optional[str] = Field(None, description="证据文件路径")


class MisstatementResponse(BaseModel):
    """错报记录响应"""
    id: str
    sample_id: str
    misstatement_type: str
    misstatement_amount: Optional[float]
    original_amount: Optional[float]
    correct_amount: Optional[float]
    description: Optional[str]
    severity: Optional[str]
    identified_by: str
    is_confirmed: bool
    created_at: datetime


@router.post("/projects/{project_id}/misstatements", response_model=MisstatementResponse, status_code=201)
async def create_misstatement(
    project_id: str,
    data: MisstatementCreate,
    current_user: UserInDB = Depends(get_current_user)
):
    """
    记录错报

    手动记录样本中发现的错报
    """
    from app.services.sample_test_service import SampleTestService

    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

    service = SampleTestService()
    ms_id = service.record_misstatement(
        sample_id=data.sample_id,
        project_id=project_id,
        misstatement_type=data.misstatement_type,
        misstatement_amount=data.misstatement_amount,
        original_amount=data.original_amount,
        correct_amount=data.correct_amount,
        description=data.description,
        evidence_path=data.evidence_path
    )

    # 返回创建的记录
    with get_db_cursor() as cursor:
        cursor.execute(
            """
            SELECT id, sample_id, misstatement_type, misstatement_amount,
                   original_amount, correct_amount, description, severity,
                   identified_by, is_confirmed, created_at
            FROM sample_misstatements WHERE id = ?
            """,
            [ms_id]
        )
        row = cursor.fetchone()

        return MisstatementResponse(
            id=row[0],
            sample_id=row[1],
            misstatement_type=row[2],
            misstatement_amount=float(row[3]) if row[3] else None,
            original_amount=float(row[4]) if row[4] else None,
            correct_amount=float(row[5]) if row[5] else None,
            description=row[6],
            severity=row[7],
            identified_by=row[8],
            is_confirmed=bool(row[9]),
            created_at=row[10]
        )


@router.get("/projects/{project_id}/misstatements", response_model=List[MisstatementResponse])
async def list_misstatements(
    project_id: str,
    current_user: UserInDB = Depends(get_current_user),
    sample_id: Optional[str] = Query(None, description="按样本筛选")
):
    """
    获取错报记录列表

    可按样本筛选或获取项目下所有错报
    """
    from app.services.sample_test_service import SampleTestService

    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

    service = SampleTestService()
    misstatements = service.get_misstatements(project_id, sample_id)

    return [
        MisstatementResponse(
            id=m["id"],
            sample_id=m["sample_id"],
            misstatement_type=m["misstatement_type"],
            misstatement_amount=m["misstatement_amount"],
            original_amount=m["original_amount"],
            correct_amount=m["correct_amount"],
            description=m["description"],
            severity=m["severity"],
            identified_by=m["identified_by"],
            is_confirmed=m["is_confirmed"],
            created_at=m["created_at"]
        )
        for m in misstatements
    ]


@router.post("/projects/{project_id}/misstatements/{misstatement_id}/evidence", status_code=201)
async def upload_misstatement_evidence(
    project_id: str,
    misstatement_id: str,
    current_user: UserInDB = Depends(get_current_user),
    file: UploadFile = File(...)
):
    """
    上传错报证据文件

    支持上传截图、文档等证据文件
    """
    with get_db_cursor() as cursor:
        # 验证项目和错报记录存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        cursor.execute(
            "SELECT id FROM sample_misstatements WHERE id = ? AND project_id = ?",
            [misstatement_id, project_id]
        )
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="错报记录不存在")

    # 创建证据存储目录
    evidence_dir = os.path.join(settings.UPLOAD_DIR, "misstatements", project_id)
    os.makedirs(evidence_dir, exist_ok=True)

    # 生成文件名
    file_ext = os.path.splitext(file.filename)[1] if file.filename else ".jpg"
    file_name = f"{misstatement_id}_{uuid.uuid4().hex[:8]}{file_ext}"
    file_path = os.path.join(evidence_dir, file_name)

    # 保存文件
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        logger.error(f"证据文件保存失败: {str(e)}")
        raise HTTPException(status_code=500, detail="文件保存失败，请稍后重试")

    # 更新错报记录的证据路径
    relative_path = f"misstatements/{project_id}/{file_name}"
    with get_db_cursor() as cursor:
        cursor.execute(
            """
            UPDATE sample_misstatements
            SET evidence_path = ?
            WHERE id = ?
            """,
            [relative_path, misstatement_id]
        )
        get_db().commit()

    return {
        "id": misstatement_id,
        "evidence_path": relative_path,
        "file_name": file.filename,
        "message": "证据上传成功"
    }


class ManualOverrideRequest(BaseModel):
    """人工修正请求"""
    ai_result_id: str = Field(..., description="AI测试结果ID")
    override_conclusion: str = Field(..., description="修正结论: valid/abnormal")
    override_reason: str = Field(..., description="修正原因")


@router.put("/projects/{project_id}/samples/{sample_id}/override")
async def manual_override_sample(
    project_id: str,
    sample_id: str,
    data: ManualOverrideRequest,
    current_user: UserInDB = Depends(get_current_user)
):
    """
    人工修正AI判定

    审计师对AI判定结果进行修正
    """
    from app.services.sample_test_service import SampleTestService

    with get_db_cursor() as cursor:
        # 验证项目和样本存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        cursor.execute("SELECT id FROM samples WHERE id = ? AND project_id = ?", [sample_id, project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="样本不存在")

    if data.override_conclusion not in ["valid", "abnormal"]:
        raise HTTPException(status_code=400, detail="修正结论必须是 valid 或 abnormal")

    service = SampleTestService()
    service.manual_override(
        sample_id=sample_id,
        ai_result_id=data.ai_result_id,
        override_conclusion=data.override_conclusion,
        override_reason=data.override_reason
    )

    return {"message": "修正成功"}


@router.get("/projects/{project_id}/test-statistics")
async def get_test_statistics(
    project_id: str,
    current_user: UserInDB = Depends(get_current_user),
    record_id: Optional[str] = Query(None, description="抽样记录ID")
):
    """
    获取测试统计

    返回样本测试的各项统计数据
    """
    from app.services.sample_test_service import SampleTestService

    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

    service = SampleTestService()
    stats = service.get_test_statistics(project_id, record_id)

    return stats


class SampleDetailResponse(BaseModel):
    """样本详情响应"""
    id: str
    voucher_id: str
    voucher_no: str
    amount: Optional[float]
    subject_name: Optional[str]
    description: Optional[str]
    risk_score: Optional[float]
    reason: Optional[str]
    test_status: Optional[str]
    ai_test_result: Optional[str]
    misstatement_amount: Optional[float]
    tested_at: Optional[datetime]
    ai_test_result_detail: Optional[dict] = None


class AITestResultDetail(BaseModel):
    """AI测试结果详情"""
    ai_conclusion: str
    confidence: float
    risk_level: Optional[str]
    risk_score: Optional[int]
    risk_factors: Optional[List[str]]
    anomaly_descriptions: Optional[List[str]]
    evidence_analysis: Optional[dict]


@router.get("/projects/{project_id}/samples/{sample_id}", response_model=SampleDetailResponse)
async def get_sample_detail(
    project_id: str,
    sample_id: str,
    current_user: UserInDB = Depends(get_current_user)
):
    """
    获取样本详情

    包含测试状态和AI判定结果详情
    """
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        cursor.execute(
            """
            SELECT s.id, s.voucher_id, v.voucher_no, v.amount, v.subject_name,
                   v.description, s.risk_score, s.reason, s.test_status,
                   s.ai_test_result, s.misstatement_amount, s.tested_at
            FROM samples s
            JOIN vouchers v ON s.voucher_id = v.id
            WHERE s.id = ? AND s.project_id = ?
            """,
            [sample_id, project_id]
        )
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="样本不存在")

        # 获取AI测试结果详情
        ai_test_result_detail = None
        cursor.execute(
            """
            SELECT ai_conclusion, confidence, risk_level, risk_factors,
                   anomaly_descriptions, evidence_analysis, risk_score
            FROM sample_ai_test_results
            WHERE sample_id = ?
            ORDER BY created_at DESC
            LIMIT 1
            """,
            [sample_id]
        )
        ai_row = cursor.fetchone()
        if ai_row:
            ai_test_result_detail = {
                "ai_conclusion": ai_row[0],
                "confidence": float(ai_row[1]) if ai_row[1] else 0.5,
                "risk_level": ai_row[2],
                "risk_factors": json.loads(ai_row[3]) if ai_row[3] else [],
                "anomaly_descriptions": json.loads(ai_row[4]) if ai_row[4] else [],
                "evidence_analysis": json.loads(ai_row[5]) if ai_row[5] else {},
                "risk_score": int(ai_row[6]) if ai_row[6] else 0
            }

        return SampleDetailResponse(
            id=row[0],
            voucher_id=row[1],
            voucher_no=row[2],
            amount=float(row[3]) if row[3] else None,
            subject_name=row[4],
            description=row[5],
            risk_score=float(row[6]) if row[6] else None,
            reason=row[7],
            test_status=row[8],
            ai_test_result=row[9],
            misstatement_amount=float(row[10]) if row[10] else None,
            tested_at=row[11],
            ai_test_result_detail=ai_test_result_detail
        )


# ==================== 总体推断API ====================

class InferenceRequest(BaseModel):
    """推断请求"""
    record_id: str = Field(..., description="抽样记录ID")
    confidence_level: float = Field(0.95, ge=0.90, le=0.99, description="置信水平")
    tolerable_error: float = Field(0.05, ge=0.01, le=0.20, description="可容忍误差率")


class InferenceResponse(BaseModel):
    """推断响应"""
    id: str
    method: str
    sample_size: int
    population_size: int
    population_amount: float
    misstatement_count: int
    misstatement_amount: float
    projected_misstatement: float
    upper_limit: float
    lower_limit: float
    confidence_level: float
    precision: float
    is_acceptable: bool
    conclusion: str
    recommendations: List[str]


class InferenceSummaryResponse(BaseModel):
    """推断摘要响应"""
    record_id: str
    rule_name: str
    rule_type: str
    sample_size: int
    population_count: int
    population_amount: float
    tested_count: int
    misstatement_count: int
    total_misstatement_amount: float
    inference_completed: bool
    inference_conclusion: Optional[str]


@router.post("/projects/{project_id}/inference", response_model=InferenceResponse)
async def execute_inference(
    project_id: str,
    request: InferenceRequest,
    current_user: UserInDB = Depends(get_current_user)
):
    """
    执行总体推断

    根据样本测试结果推断总体错报
    """
    from app.services.inference_service import InferenceService

    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        # 验证抽样记录存在
        cursor.execute(
            "SELECT id FROM sampling_records WHERE id = ? AND project_id = ?",
            [request.record_id, project_id]
        )
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="抽样记录不存在")

    service = InferenceService()
    result = service.infer(
        project_id=project_id,
        record_id=request.record_id,
        confidence_level=request.confidence_level,
        tolerable_error=request.tolerable_error
    )

    # 获取保存的推断ID
    with get_db_cursor() as cursor:
        cursor.execute(
            """
            SELECT id FROM statistical_inferences
            WHERE record_id = ?
            ORDER BY calculated_at DESC
            LIMIT 1
            """,
            [request.record_id]
        )
        row = cursor.fetchone()
        inference_id = row[0] if row else ""

    return InferenceResponse(
        id=inference_id,
        method=result.method.value,
        sample_size=result.sample_size,
        population_size=result.population_size,
        population_amount=result.population_amount,
        misstatement_count=result.misstatement_count,
        misstatement_amount=result.misstatement_amount,
        projected_misstatement=result.projected_misstatement,
        upper_limit=result.upper_limit,
        lower_limit=result.lower_limit,
        confidence_level=result.confidence_level,
        precision=result.precision,
        is_acceptable=result.is_acceptable,
        conclusion=result.conclusion,
        recommendations=result.recommendations
    )


@router.get("/projects/{project_id}/inference/{inference_id}", response_model=InferenceResponse)
async def get_inference_result(
    project_id: str,
    inference_id: str,
    current_user: UserInDB = Depends(get_current_user)
):
    """
    获取推断结果

    根据ID获取已保存的推断结果
    """
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        cursor.execute(
            """
            SELECT id, inference_type, sample_misstatement_count, sample_misstatement_amount,
                   projected_misstatement, upper_limit, lower_limit, confidence_level,
                   precision, conclusion, is_acceptable, recommendations
            FROM statistical_inferences
            WHERE id = ? AND project_id = ?
            """,
            [inference_id, project_id]
        )
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="推断结果不存在")

        return InferenceResponse(
            id=row[0],
            method=row[1],
            sample_size=0,  # 需要从关联记录获取
            population_size=0,
            population_amount=0,
            misstatement_count=row[2],
            misstatement_amount=float(row[3]) if row[3] else 0,
            projected_misstatement=float(row[4]) if row[4] else 0,
            upper_limit=float(row[5]) if row[5] else 0,
            lower_limit=float(row[6]) if row[6] else 0,
            confidence_level=float(row[7]) if row[7] else 0.95,
            precision=float(row[8]) if row[8] else 0,
            conclusion=row[9],
            is_acceptable=bool(row[10]),
            recommendations=json.loads(row[11]) if isinstance(row[11], str) else row[11]
        )


@router.get("/projects/{project_id}/inference-summary", response_model=List[InferenceSummaryResponse])
async def get_inference_summaries(
    project_id: str,
    current_user: UserInDB = Depends(get_current_user)
):
    """
    获取推断摘要列表

    返回项目下所有抽样记录的推断摘要
    """
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        cursor.execute(
            """
            SELECT r.id, r.rule_name, r.rule_type, r.sample_size, r.population_count,
                   r.population_amount, r.tested_count, r.misstatement_count,
                   r.total_misstatement_amount, r.inference_completed, r.inference_conclusion
            FROM sampling_records r
            WHERE r.project_id = ?
            ORDER BY r.created_at DESC
            """,
            [project_id]
        )
        rows = cursor.fetchall()

        return [
            InferenceSummaryResponse(
                record_id=row[0],
                rule_name=row[1],
                rule_type=row[2],
                sample_size=row[3] or 0,
                population_count=row[4] or 0,
                population_amount=float(row[5]) if row[5] else 0,
                tested_count=row[6] or 0,
                misstatement_count=row[7] or 0,
                total_misstatement_amount=float(row[8]) if row[8] else 0,
                inference_completed=bool(row[9]),
                inference_conclusion=row[10]
            )
            for row in rows
        ]


# ==================== 抽样策略API ====================

from app.services.sampling_strategy_service import sampling_strategy_recommender, SamplingMethod
from app.services.risk_profile_service import risk_profile_generator, RiskLevel


class SampleSizeRequest(BaseModel):
    """样本量计算请求"""
    population_size: int = Field(..., description="总体规模")
    confidence_level: float = Field(0.95, description="置信水平 (0.90, 0.95, 0.99)")
    tolerable_error: float = Field(0.05, description="可容忍误差率")
    expected_error: float = Field(0.03, description="预期偏差率")


class SampleSizeResponse(BaseModel):
    """样本量计算响应"""
    sample_size: int
    sampling_rate: float
    parameters: Dict[str, Any]
    curve_data: List[Dict[str, Any]]
    comparisons: List[Dict[str, Any]]


class StrategyResponse(BaseModel):
    """抽样策略响应"""
    project_id: str
    subject_code: str
    subject_name: str
    risk_level: str
    risk_score: float
    method: str
    sample_size: int
    confidence_level: float
    tolerable_error: float
    expected_error: float
    rationale: str
    total_population: int
    sampling_rate: float
    key_focus_areas: List[str]
    stratification: List[Dict[str, Any]] = []


@router.post("/projects/{project_id}/sample-size", response_model=SampleSizeResponse)
async def calculate_sample_size(
    project_id: str,
    request: SampleSizeRequest,
    current_user: UserInDB = Depends(get_current_user)
):
    """
    计算样本量

    使用统计公式计算推荐样本量
    """
    # 验证项目存在
    with get_db_cursor() as cursor:
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

    result = sampling_strategy_recommender.calculate_sample_size_interactive(
        population_size=request.population_size,
        confidence_level=request.confidence_level,
        tolerable_error=request.tolerable_error,
        expected_error=request.expected_error
    )

    return SampleSizeResponse(
        sample_size=result["sample_size"],
        sampling_rate=result["sampling_rate"],
        parameters=result["parameters"],
        curve_data=result["curve_data"],
        comparisons=result["comparisons"]
    )


@router.get("/projects/{project_id}/population-stats")
async def get_population_stats(
    project_id: str,
    current_user: UserInDB = Depends(get_current_user),
    subject_code: Optional[str] = Query(None, description="科目代码过滤")
):
    """
    获取总体统计信息

    返回项目凭证的总体规模、金额分布等统计信息
    """
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

        # 主统计查询
        if subject_code:
            cursor.execute(
                """
                SELECT
                    COUNT(*) as count,
                    COALESCE(SUM(amount), 0) as total_amount,
                    COALESCE(AVG(amount), 0) as avg_amount,
                    COALESCE(MIN(amount), 0) as min_amount,
                    COALESCE(MAX(amount), 0) as max_amount
                FROM vouchers
                WHERE project_id = ? AND subject_code LIKE ?
                """,
                [project_id, f"{subject_code}%"]
            )
        else:
            cursor.execute(
                """
                SELECT
                    COUNT(*) as count,
                    COALESCE(SUM(amount), 0) as total_amount,
                    COALESCE(AVG(amount), 0) as avg_amount,
                    COALESCE(MIN(amount), 0) as min_amount,
                    COALESCE(MAX(amount), 0) as max_amount
                FROM vouchers
                WHERE project_id = ?
                """,
                [project_id]
            )
        row = cursor.fetchone()

        # 获取科目列表
        cursor.execute(
            """
            SELECT DISTINCT subject_code, subject_name, COUNT(*) as count
            FROM vouchers
            WHERE project_id = ? AND subject_code IS NOT NULL
            GROUP BY subject_code, subject_name
            ORDER BY count DESC
            """,
            [project_id]
        )
        subjects = [
            {"code": r[0], "name": r[1], "count": r[2]}
            for r in cursor.fetchall()
        ]

    return {
        "population_size": row[0] or 0,
        "total_amount": float(row[1]) if row[1] else 0,
        "avg_amount": float(row[2]) if row[2] else 0,
        "min_amount": float(row[3]) if row[3] else 0,
        "max_amount": float(row[4]) if row[4] else 0,
        "subjects": subjects
    }


@router.get("/projects/{project_id}/strategy/recommend")
async def recommend_sampling_strategy(
    project_id: str,
    current_user: UserInDB = Depends(get_current_user),
    subject_code: Optional[str] = Query(None, description="科目代码（不传则对整个项目推荐）"),
    confidence_level: float = Query(0.95, description="置信水平"),
    tolerable_error: Optional[float] = Query(None, description="可容忍误差率"),
    expected_error: Optional[float] = Query(None, description="预期偏差率")
):
    """
    推荐抽样策略

    根据科目风险画像自动推荐抽样方法、样本量和策略
    如果不传科目代码，则对整个项目生成综合策略推荐
    """
    with get_db_cursor() as cursor:
        # 验证项目存在
        cursor.execute("SELECT id FROM projects WHERE id = ?", [project_id])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="项目不存在")

    # 项目级别的策略推荐
    if not subject_code:
        return await _recommend_project_strategy(
            project_id, confidence_level, tolerable_error, expected_error
        )

    # 科目级别的策略推荐
    risk_profile = risk_profile_generator.get_profile(project_id, subject_code)

    if not risk_profile:
        # 如果没有风险画像，先生成一个
        risk_profile = risk_profile_generator.generate(
            project_id=project_id,
            subject_code=subject_code
        )
        risk_profile_generator.save_profile(risk_profile)

    # 推荐抽样策略
    recommendation = sampling_strategy_recommender.recommend(
        risk_profile=risk_profile,
        confidence_level=confidence_level,
        tolerable_error=tolerable_error,
        expected_error=expected_error
    )

    return StrategyResponse(
        project_id=recommendation.project_id,
        subject_code=recommendation.subject_code,
        subject_name=recommendation.subject_name,
        risk_level=recommendation.risk_level.value,
        risk_score=recommendation.risk_score,
        method=recommendation.strategy.method.value,
        sample_size=recommendation.strategy.sample_size,
        confidence_level=recommendation.strategy.confidence_level,
        tolerable_error=recommendation.strategy.tolerable_error,
        expected_error=recommendation.strategy.expected_error,
        rationale=recommendation.strategy.rationale,
        total_population=recommendation.total_population,
        sampling_rate=recommendation.sampling_rate,
        key_focus_areas=recommendation.key_focus_areas,
        stratification=[
            {
                "name": l.name,
                "min_amount": l.min_amount,
                "max_amount": l.max_amount,
                "count": l.count,
                "sample_size": l.sample_size,
                "method": l.method.value
            }
            for l in recommendation.stratification
        ]
    )


async def _recommend_project_strategy(
    project_id: str,
    confidence_level: float = 0.95,
    tolerable_error: Optional[float] = None,
    expected_error: Optional[float] = None
):
    """
    项目级别的综合策略推荐

    对整个项目的所有凭证进行分析，生成综合抽样策略
    """
    # 获取项目总体统计
    with get_db_cursor() as cursor:
        # 获取总体统计
        cursor.execute(
            """
            SELECT
                COUNT(*) as count,
                COALESCE(SUM(amount), 0) as total_amount,
                COALESCE(AVG(amount), 0) as avg_amount,
                COALESCE(MIN(amount), 0) as min_amount,
                COALESCE(MAX(amount), 0) as max_amount
            FROM vouchers
            WHERE project_id = ?
            """,
            [project_id]
        )
        row = cursor.fetchone()
        population_size = row[0] or 0
        total_amount = float(row[1]) if row[1] else 0
        avg_amount = float(row[2]) if row[2] else 0
        min_amount = float(row[3]) if row[3] else 0
        max_amount = float(row[4]) if row[4] else 0

        # 获取各科目统计
        cursor.execute(
            """
            SELECT
                subject_code,
                subject_name,
                COUNT(*) as count,
                SUM(amount) as total_amt
            FROM vouchers
            WHERE project_id = ? AND subject_code IS NOT NULL
            GROUP BY subject_code, subject_name
            ORDER BY total_amt DESC
            """,
            [project_id]
        )
        subjects_data = cursor.fetchall()

        # 获取风险画像汇总
        cursor.execute(
            """
            SELECT
                risk_level,
                COUNT(*) as count,
                AVG(anomaly_score) as avg_score
            FROM risk_profiles
            WHERE project_id = ?
            GROUP BY risk_level
            """,
            [project_id]
        )
        risk_summary = cursor.fetchall()

    if population_size == 0:
        raise HTTPException(status_code=400, detail="项目中没有凭证数据")

    # 设置默认参数
    if tolerable_error is None:
        tolerable_error = 0.05
    if expected_error is None:
        expected_error = 0.03

    # 根据风险画像汇总确定整体风险等级
    risk_level = "medium"  # 默认中风险
    risk_score = 50.0

    if risk_summary:
        high_count = 0
        medium_count = 0
        low_count = 0
        total_risk_score = 0.0
        total_count = 0

        for r in risk_summary:
            level, count, avg_score = r[0], r[1], r[2] or 0
            total_count += count
            total_risk_score += avg_score * count
            if level == "high":
                high_count = count
            elif level == "medium":
                medium_count = count
            else:
                low_count = count

        if total_count > 0:
            risk_score = total_risk_score / total_count

        # 确定整体风险等级
        if high_count > medium_count and high_count > low_count:
            risk_level = "high"
        elif low_count > high_count and low_count > medium_count:
            risk_level = "low"
        else:
            risk_level = "medium"

    # 计算样本量
    z = {0.90: 1.645, 0.95: 1.96, 0.99: 2.576}.get(confidence_level, 1.96)
    n = (z ** 2 * expected_error * (1 - expected_error)) / (tolerable_error ** 2)
    n_adjusted = n * population_size / (population_size + n - 1)
    sample_size = int(round(n_adjusted))
    sample_size = min(sample_size, population_size)
    sample_size = max(1, sample_size)

    # 根据风险等级确定抽样方法
    if risk_level == "high":
        method = "stratified"
        method_label = "分层抽样"
        # 高风险项目：分层抽样，大额凭证重点审查
        stratification = _create_project_stratification(
            project_id, population_size, sample_size, min_amount, max_amount
        )
    elif risk_level == "low":
        method = "random"
        method_label = "随机抽样"
        stratification = []
    else:
        method = "stratified"
        method_label = "分层抽样"
        stratification = _create_project_stratification(
            project_id, population_size, sample_size, min_amount, max_amount
        )

    sampling_rate = sample_size / population_size if population_size > 0 else 0

    # 生成策略依据
    rationale = f"该项目共有 {population_size} 笔凭证，总金额 {total_amount:,.2f} 元。"
    rationale += f"根据风险画像分析，整体风险等级为【{'高' if risk_level == 'high' else '中' if risk_level == 'medium' else '低'}风险】，"
    rationale += f"综合风险分数 {risk_score:.1f} 分。"
    rationale += f"建议采用{method_label}方法，在 {confidence_level*100:.0f}% 置信水平下，抽取 {sample_size} 笔样本，抽样比例 {sampling_rate*100:.1f}%。"

    # 确定重点关注领域
    key_focus_areas = []
    if risk_level == "high":
        key_focus_areas = ["大额交易核实", "异常凭证审查", "关联方交易识别", "截止测试"]
    elif risk_level == "medium":
        key_focus_areas = ["大额交易核实", "抽样检查", "异常凭证审查"]

    return {
        "project_id": project_id,
        "subject_code": "ALL",
        "subject_name": "项目整体",
        "risk_level": risk_level,
        "risk_score": risk_score,
        "method": method,
        "sample_size": sample_size,
        "confidence_level": confidence_level,
        "tolerable_error": tolerable_error,
        "expected_error": expected_error,
        "rationale": rationale,
        "total_population": population_size,
        "sampling_rate": sampling_rate,
        "key_focus_areas": key_focus_areas,
        "stratification": stratification,
        "subjects_analysis": [
            {
                "code": s[0],
                "name": s[1],
                "count": s[2],
                "total_amount": float(s[3]) if s[3] else 0
            }
            for s in subjects_data[:10]  # 返回前10个科目
        ]
    }


def _create_project_stratification(
    project_id: str,
    population_size: int,
    total_sample: int,
    min_amount: float,
    max_amount: float
) -> List[Dict[str, Any]]:
    """创建项目级别的金额分层"""
    with get_db_cursor() as cursor:
        # 获取金额分布
        cursor.execute(
            """
            SELECT amount FROM vouchers
            WHERE project_id = ? AND amount IS NOT NULL
            ORDER BY amount DESC
            """,
            [project_id]
        )
        amounts = [float(row[0]) for row in cursor.fetchall()]

    if not amounts:
        return []

    # 计算分层阈值
    total_amt = sum(amounts)
    threshold_80 = total_amt * 0.8  # 前80%金额阈值
    threshold_95 = total_amt * 0.95  # 前95%金额阈值

    large_threshold = 0
    medium_threshold = 0
    cumulative = 0

    for amount in amounts:
        cumulative += amount
        if large_threshold == 0 and cumulative >= threshold_80:
            large_threshold = amount
        if medium_threshold == 0 and cumulative >= threshold_95:
            medium_threshold = amount
            break

    # 统计各层数量和样本分配
    large_count = len([a for a in amounts if a >= large_threshold]) if large_threshold > 0 else 0
    medium_count = len([a for a in amounts if medium_threshold <= a < large_threshold]) if medium_threshold > 0 else 0
    small_count = len(amounts) - large_count - medium_count

    # 样本分配：大额50%，中额30%，小额20%
    large_sample = int(total_sample * 0.5)
    medium_sample = int(total_sample * 0.3)
    small_sample = total_sample - large_sample - medium_sample

    layers = []

    if large_count > 0:
        layers.append({
            "name": "大额层",
            "min_amount": large_threshold,
            "max_amount": max_amount,
            "count": large_count,
            "sample_size": min(large_sample, large_count),
            "method": "judgment"
        })

    if medium_count > 0:
        layers.append({
            "name": "中额层",
            "min_amount": medium_threshold if medium_threshold > 0 else large_threshold,
            "max_amount": large_threshold if large_threshold > 0 else max_amount,
            "count": medium_count,
            "sample_size": min(medium_sample, medium_count),
            "method": "random"
        })

    if small_count > 0:
        layers.append({
            "name": "小额层",
            "min_amount": min_amount,
            "max_amount": medium_threshold if medium_threshold > 0 else large_threshold,
            "count": small_count,
            "sample_size": min(small_sample, small_count),
            "method": "random"
        })

    return layers